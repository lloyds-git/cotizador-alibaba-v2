"""
Ingest de productos desde xlsx intermedios (los que produce pdf_a_formato_hd.py
en el paso 3) a la BD SQLite.

Estrategia:
- Idempotente: si (proveedor, sku) ya existe, actualiza datos en vez de duplicar.
- Copia las imagenes desde donde esten al directorio data/fotos/.
- No toca el campo marcado_cotizar al actualizar (preserva eleccion del usuario).
"""

from __future__ import annotations

import json
import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl
from sqlalchemy.orm import Session

from app.modelos import Proveedor, Producto, Foto


# Sufijos / palabras corporativas que aparecen variablemente cuando Claude
# extrae el seller del PDF (a veces con LTD, a veces sin; con o sin coma,
# con "Trade", "Trading", "International", etc.). Se removen al normalizar
# para que "Shenzhen Dongtai Sponge Product CO., LTD" matchee con
# "Shenzhen Dongtai Sponge Product CO." y no se duplique el proveedor.
_PALABRAS_CORPORATIVAS = (
    "co", "corp", "ltd", "inc", "llc", "company", "corporation", "limited",
    "trade", "trading", "import", "imports", "export", "exports",
    "international", "intl", "group", "enterprise", "enterprises",
    "industry", "industrial", "industries", "manufacturing", "mfg",
    "products", "product", "technology", "tech", "material", "materials",
)
# Patron precompilado: word boundary + palabra + opcional punto/coma
_RE_PALABRAS = re.compile(
    r"\b(" + "|".join(re.escape(w) for w in _PALABRAS_CORPORATIVAS) + r")\b\.?,?",
    re.IGNORECASE,
)
_RE_PUNTUACION = re.compile(r"[,.()\-&]+")
_RE_WS = re.compile(r"\s+")


def normalizar_nombre_proveedor(nombre: str) -> str:
    """Normaliza un nombre de proveedor para hacer matching tolerante.

    Quita sufijos corporativos (Co., Ltd, Trade Co., Inc, etc.), puntuacion
    y normaliza espacios + case. Resultado deterministico, sin fuzzy
    matching: dos nombres con la misma forma normalizada matchean.

    Ejemplos:
      "Shenzhen Dongtai Sponge Product CO., LTD" -> "shenzhen dongtai sponge"
      "Shenzhen Dongtai Sponge Product CO."       -> "shenzhen dongtai sponge"
      "ZHANGJIAGANG KINGTALE INTL TRADING CO.,LTD"-> "zhangjiagang kingtale"
    """
    if not nombre:
        return ""
    s = nombre.lower()
    s = _RE_PUNTUACION.sub(" ", s)
    s = _RE_PALABRAS.sub(" ", s)
    s = _RE_WS.sub(" ", s).strip()
    return s


def buscar_proveedor_existente(
    session: Session, nombre: str
) -> Proveedor | None:
    """Busca un proveedor existente con match tolerante.

    Estrategia:
      1. Match exacto por nombre (rapido, indice implicito).
      2. Si no, normaliza el candidato y compara contra todos los
         proveedores en BD por su forma normalizada.

    Devuelve None si ningun proveedor coincide tras normalizar.
    """
    prov = session.query(Proveedor).filter_by(nombre=nombre).first()
    if prov:
        return prov
    nombre_norm = normalizar_nombre_proveedor(nombre)
    if not nombre_norm:
        return None
    for p in session.query(Proveedor).all():
        if normalizar_nombre_proveedor(p.nombre) == nombre_norm:
            return p
    return None


def _cargar_meta(xlsx_path: str) -> dict:
    """Lee el .meta.json junto al xlsx intermedio. Devuelve {} si no existe o falla."""
    meta_path = Path(xlsx_path + ".meta.json")
    if not meta_path.exists():
        return {}
    try:
        return json.loads(meta_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {}


def resolver_nombre_proveedor(xlsx_path: str, fallback: str | None = None) -> str:
    """
    Resuelve el nombre del proveedor en este orden:
      1. .meta.json junto al xlsx (key 'seller', extraido del PDF con Claude).
      2. fallback si se paso.
      3. Nombre del archivo limpio (legacy).
    """
    seller = (_cargar_meta(xlsx_path).get("seller") or "").strip()
    if seller:
        return seller[:200]
    if fallback:
        return fallback
    return Path(xlsx_path).stem.replace("_intermedio_", "").replace("_", " ")[:60]


def resolver_archivo_pdf(xlsx_path: str) -> str:
    """
    Resuelve el nombre del PDF original del que vino este intermedio.
      1. .meta.json key 'pdf_original'.
      2. Fallback: el nombre del .xlsx intermedio (legacy, lo que se guardaba antes).
    """
    pdf_original = (_cargar_meta(xlsx_path).get("pdf_original") or "").strip()
    if pdf_original:
        return pdf_original
    return Path(xlsx_path).name


# Columnas esperadas en el xlsx intermedio
COL_FOTO = 1
COL_SKU = 2
COL_DESC = 3
COL_MEDIDAS = 4
COL_MATERIAL = 5
COL_PESO = 6
COL_COLOR = 7
COL_MOQ = 8
COL_PACKING = 9
COL_CARTON = 10
COL_CBM = 11
COL_PZAS20 = 12
COL_PZAS40 = 13
COL_LEAD = 14
COL_FOB = 15


def _to_float(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(str(v).strip())
    except (ValueError, TypeError):
        return None


def _to_int(v) -> int | None:
    f = _to_float(v)
    return int(f) if f is not None else None


def _extraer_imagenes_xlsx(xlsx_path: str) -> dict[int, bytes]:
    """
    Devuelve {fila_excel: bytes_de_imagen} para imagenes ancladas en col A.
    Lee el ZIP del xlsx directamente para obtener bytes.
    """
    resultado: dict[int, bytes] = {}
    with zipfile.ZipFile(xlsx_path) as z:
        nombres = set(z.namelist())
        if "xl/drawings/drawing1.xml" not in nombres:
            return {}
        if "xl/drawings/_rels/drawing1.xml.rels" not in nombres:
            return {}

        rels_xml = z.read("xl/drawings/_rels/drawing1.xml.rels").decode("utf-8")
        ns_rels = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
        rid_a_target = {}
        for rel in ET.fromstring(rels_xml).findall("r:Relationship", ns_rels):
            rid_a_target[rel.get("Id")] = rel.get("Target")

        drawing_xml = z.read("xl/drawings/drawing1.xml").decode("utf-8")
        ns = {
            "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
            "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
            "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        }
        root = ET.fromstring(drawing_xml)
        anchors = []
        for tag in ("oneCellAnchor", "twoCellAnchor"):
            anchors.extend(root.findall(f"xdr:{tag}", ns))

        for a in anchors:
            from_node = a.find("xdr:from", ns)
            pic = a.find("xdr:pic", ns)
            if from_node is None or pic is None:
                continue
            col = int(from_node.find("xdr:col", ns).text)
            if col != 0:
                continue
            fila = int(from_node.find("xdr:row", ns).text) + 1
            blip = pic.find("xdr:blipFill/a:blip", ns)
            rid = blip.get(
                "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed"
            )
            target = rid_a_target.get(rid, "")
            if target.startswith("/"):
                ruta_zip = target.lstrip("/")
            else:
                import os
                ruta_zip = os.path.normpath(
                    os.path.join("xl/drawings", target)
                ).replace("\\", "/")
            if ruta_zip in nombres:
                resultado[fila] = z.read(ruta_zip)
    return resultado


def ingestar_xlsx_intermedio(
    session: Session,
    xlsx_path: str,
    nombre_proveedor: str | None,
    fotos_destino: str,
) -> int:
    """
    Lee un xlsx intermedio (el output de pdf_a_formato_hd.py paso 3) y lo
    inserta/actualiza en la BD.

    Si nombre_proveedor es None, se resuelve via resolver_nombre_proveedor
    (lee .meta.json o cae al nombre del archivo).

    Devuelve el numero de productos NUEVOS insertados (no incluye actualizados).
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    nombre_proveedor = resolver_nombre_proveedor(xlsx_path, fallback=nombre_proveedor)
    archivo_pdf = resolver_archivo_pdf(xlsx_path)

    prov = buscar_proveedor_existente(session, nombre_proveedor)
    if prov is None:
        prov = Proveedor(
            nombre=nombre_proveedor,
            archivo_pdf=archivo_pdf,
        )
        session.add(prov)
        session.flush()
    elif prov.archivo_pdf != archivo_pdf:
        # Mantener actualizado el path al PDF original (puede cambiar entre re-ingestas)
        prov.archivo_pdf = archivo_pdf

    imagenes = _extraer_imagenes_xlsx(xlsx_path)

    fotos_dir = Path(fotos_destino)
    fotos_dir.mkdir(parents=True, exist_ok=True)

    import hashlib

    # Tracking de SKUs ya usados EN ESTE xlsx para detectar colisiones
    # entre filas (caso comun: proveedor lista 3 tiers de pricing con el
    # mismo SKU real, ej. PD0421 con MOQ 500/1000/2000 y FOB 5.6/5.2/5.0).
    # Sin esto, el UNIQUE(proveedor_id, sku) hace UPDATE en cada colision
    # y solo queda la ultima fila.
    skus_usados: dict[str, int] = {}

    nuevos = 0
    for fila in range(2, ws.max_row + 1):
        sku = ws.cell(fila, COL_SKU).value or ""
        sku = str(sku).strip()
        desc = ws.cell(fila, COL_DESC).value or ""
        desc = str(desc).strip()
        if not (sku or desc):
            continue

        # Leemos los campos diferenciadores ahora para poder usarlos en
        # el hash del SKU sintetico si no hay SKU real.
        medidas = str(ws.cell(fila, COL_MEDIDAS).value or "").strip()
        material = str(ws.cell(fila, COL_MATERIAL).value or "").strip()
        fob_usd = _to_float(ws.cell(fila, COL_FOB).value)

        # Si no hay SKU real, generamos uno sintetico determinista. El hash
        # incluye desc + medidas + material + fob para que VARIANTES con el
        # mismo nombre (ej. "PET Carrier" 10kg vs 15kg con distinta talla y
        # precio) no colisionen al mismo AUTO-{hash} y se sobreescriban por
        # el UNIQUE(proveedor_id, sku). Sigue siendo deterministico:
        # re-ingestar el mismo xlsx produce los mismos AUTO-* y actualiza
        # en vez de duplicar.
        sku_sintetico = False
        if not sku:
            clave = "|".join([desc, medidas, material, f"{fob_usd if fob_usd is not None else ''}"])
            h = hashlib.md5(clave.encode("utf-8")).hexdigest()[:10]
            sku = f"AUTO-{h}"
            sku_sintetico = True

        # Si el SKU ya fue usado en una fila previa del MISMO xlsx, no es
        # un duplicado real (re-ingest del mismo archivo) sino una variante
        # con el mismo codigo: suffixar con hash de campos diferenciadores.
        # Hash deterministico => idempotente entre re-ingests. La primera
        # ocurrencia mantiene el SKU base; las subsecuentes reciben sufijo.
        if sku in skus_usados:
            clave_var = "|".join(
                [desc, medidas, material, f"{fob_usd if fob_usd is not None else ''}"]
            )
            h_var = hashlib.md5(clave_var.encode("utf-8")).hexdigest()[:6]
            sku_propuesto = f"{sku}-{h_var}"
            # Fallback si dos filas son TRULY identicas (mismo desc+medidas
            # +material+fob): contador secuencial para no chocar.
            contador = 2
            while sku_propuesto in skus_usados:
                sku_propuesto = f"{sku}-{h_var}-V{contador}"
                contador += 1
            sku = sku_propuesto
        skus_usados[sku] = skus_usados.get(sku, 0) + 1

        prod = session.query(Producto).filter_by(
            proveedor_id=prov.id, sku=sku
        ).first()

        es_nuevo = prod is None
        if es_nuevo:
            prod = Producto(proveedor_id=prov.id, sku=sku, descripcion=desc)
            session.add(prod)

        prod.descripcion = desc
        prod.fob_usd = fob_usd
        prod.medidas = medidas
        prod.material = material
        prod.peso_kg = _to_float(ws.cell(fila, COL_PESO).value)
        prod.color = str(ws.cell(fila, COL_COLOR).value or "").strip()
        prod.moq = str(ws.cell(fila, COL_MOQ).value or "").strip()
        prod.packing = str(ws.cell(fila, COL_PACKING).value or "").strip()
        prod.carton_dims = str(ws.cell(fila, COL_CARTON).value or "").strip()
        prod.cbm = _to_float(ws.cell(fila, COL_CBM).value)
        prod.pzas_20ft = _to_int(ws.cell(fila, COL_PZAS20).value)
        prod.pzas_40hq = _to_int(ws.cell(fila, COL_PZAS40).value)
        prod.lead_time = str(ws.cell(fila, COL_LEAD).value or "").strip()

        session.flush()

        if fila in imagenes and es_nuevo:
            data = imagenes[fila]
            ext = ".png"
            if data[:3] == b"\xff\xd8\xff":
                ext = ".jpg"
            nombre_archivo = f"{prov.id}_{prod.id}_{sku or fila}{ext}"
            nombre_archivo = nombre_archivo.replace("/", "_").replace("\\", "_")
            destino = fotos_dir / nombre_archivo
            destino.write_bytes(data)

            foto = Foto(
                producto_id=prod.id,
                ruta_relativa=f"fotos/{nombre_archivo}",
                es_principal=True,
            )
            session.add(foto)

        if es_nuevo:
            nuevos += 1

    return nuevos
