"""Llenado del formato HD con openpyxl (cross-platform, sin Excel/COM).

Version Linux/servidor del llenado que en Windows hace `llenar_formato_hd.py`
via Excel COM (win32com). Produce el MISMO resultado: mismo layout transpuesto
(productos en columnas C, D, E...; campos en filas 2-18), mismas CONSTANTES y
mismo mapeo, para que el export HD sea consistente en cualquier entorno.

Piezas:
  - extraer_imagenes_drawing / _in_cell: leen las fotos del xlsx intermedio
    directamente del ZIP (Python puro, ya usado por el script Windows).
  - convertir_a_xlsx: xlsb -> xlsx via LibreOffice headless (openpyxl no lee
    .xlsb). Cachea el resultado junto al original.
  - asegurar_template_xlsx: devuelve un template .xlsx, convirtiendo si hace
    falta.
  - llenar_formato_hd: rellena una copia del template y la guarda como xlsx.

El intermedio de app/exportar.py tiene 18 columnas (A=foto, B=SKU, C=Descripcion,
..., O=Venta HD, P=Retail, Q=Margen, R=Proveedor) con datos desde la fila 2.
"""
from __future__ import annotations

import os
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU

# Constantes fijas por columna de producto (fila_destino -> valor). Actuan como
# fallback: el mapeo de datos solo sobrescribe si el origen trae un valor no vacio.
# Debe coincidir con llenar_formato_hd.CONSTANTES (paridad Windows/Linux).
CONSTANTES: dict[int, str] = {
    3: "Totikay Pets SA de CV",
    4: "TBD",
    5: "TBD",
    6: "PRIMARY",
    7: "TBD",
    10: "China",
    12: "NA",
    13: "NA",
    14: "NA",
    15: "NA",
    18: "NA",
}

# Mapeo por defecto: columna del origen (numero) -> fila del destino.
MAPEO_DEFAULT: dict[int, int] = {3: 8, 15: 11, 21: 16, 22: 17}

# Filas del destino con formato numerico (paridad con el script Windows).
FMT_MONEDA_FILAS = (11, 16)
FMT_PORCENTAJE_FILAS = (17,)
FMT_MONEDA = "$#,##0.00"
FMT_PORCENTAJE = "0.00%"

# Geometria de la fila de foto (igual que el script Windows: ~270x190 px).
FOTO_FILA = 2
ANCHO_COL_CHARS = 34.45     # ancho de las columnas de producto (como col C del template)
ALTO_FILA_FOTO_PT = 120.0   # alto de la fila 2 (como el template)


def col_letra_a_num(letra: str) -> int:
    """'A' -> 1, 'Z' -> 26, 'AA' -> 27."""
    letra = letra.strip().upper()
    n = 0
    for ch in letra:
        if not ("A" <= ch <= "Z"):
            raise ValueError(f"Letra de columna invalida: {letra!r}")
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def parsear_mapeo(s: str) -> dict[int, int]:
    """Parsea 'B=5,C=8,O=11' a {col_num: fila_dest}."""
    resultado: dict[int, int] = {}
    for parte in s.split(","):
        parte = parte.strip()
        if not parte:
            continue
        if "=" not in parte:
            raise ValueError(f"Mapeo invalido (falta '='): {parte!r}")
        col, fila = parte.split("=", 1)
        resultado[col_letra_a_num(col)] = int(fila.strip())
    return resultado


# --------------------------------------------------------------------------
# Extraccion de imagenes del xlsx intermedio (Python puro; copiado de
# llenar_formato_hd.py para no depender de ese modulo, que importa win32com).
# --------------------------------------------------------------------------

def extraer_imagenes_drawing(archivo_xlsx: str, tmpdir: str) -> dict[int, str]:
    """Extrae imagenes flotantes (Pictures/Shapes) ancladas en la columna A.

    Devuelve {fila_excel: ruta_local_a_imagen}.
    """
    if not archivo_xlsx.lower().endswith(".xlsx"):
        return {}
    resultado: dict[int, str] = {}
    try:
        with zipfile.ZipFile(archivo_xlsx) as z:
            nombres = set(z.namelist())
            if "xl/drawings/drawing1.xml" not in nombres:
                return {}
            if "xl/drawings/_rels/drawing1.xml.rels" not in nombres:
                return {}
            rels_xml = z.read("xl/drawings/_rels/drawing1.xml.rels").decode("utf-8")
            ns_rels = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
            rels_root = ET.fromstring(rels_xml)
            rid_a_target: dict[str, str] = {}
            for rel in rels_root.findall("r:Relationship", ns_rels):
                rid = rel.get("Id")
                target = rel.get("Target")
                if rid and target:
                    rid_a_target[rid] = target

            drawing_xml = z.read("xl/drawings/drawing1.xml").decode("utf-8")
            ns = {
                "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
                "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
                "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
            }
            root = ET.fromstring(drawing_xml)
            anchors = []
            for tag in ("oneCellAnchor", "twoCellAnchor", "absoluteAnchor"):
                anchors.extend(root.findall(f"xdr:{tag}", ns))

            for anchor in anchors:
                pic = anchor.find("xdr:pic", ns)
                if pic is None:
                    continue
                from_node = anchor.find("xdr:from", ns)
                if from_node is None:
                    continue
                col_node = from_node.find("xdr:col", ns)
                row_node = from_node.find("xdr:row", ns)
                if col_node is None or row_node is None:
                    continue
                try:
                    col_0 = int(col_node.text)
                    row_0 = int(row_node.text)
                except (TypeError, ValueError):
                    continue
                if col_0 != 0:  # solo columna A
                    continue
                fila_excel = row_0 + 1
                if fila_excel in resultado:
                    continue
                blip = pic.find("xdr:blipFill/a:blip", ns)
                if blip is None:
                    continue
                rid = blip.get(
                    "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed"
                )
                if not rid:
                    continue
                target = rid_a_target.get(rid)
                if not target:
                    continue
                if target.startswith("/"):
                    ruta_zip = target.lstrip("/")
                else:
                    ruta_zip = os.path.normpath(
                        os.path.join("xl/drawings", target)
                    ).replace("\\", "/")
                if ruta_zip not in nombres:
                    continue
                ext = os.path.splitext(ruta_zip)[1] or ".png"
                ruta_local = os.path.join(tmpdir, f"flotante_fila{fila_excel}{ext}")
                with z.open(ruta_zip) as src, open(ruta_local, "wb") as dst:
                    dst.write(src.read())
                resultado[fila_excel] = ruta_local
    except Exception as e:
        print(f"  Aviso: no se pudieron extraer imagenes flotantes ({e})")
        return {}
    return resultado


def extraer_imagenes_in_cell(archivo_xlsx: str, tmpdir: str) -> dict[int, str]:
    """Extrae imagenes 'in-cell' (Excel 365) de la columna A. {fila: ruta}."""
    if not archivo_xlsx.lower().endswith(".xlsx"):
        return {}
    resultado: dict[int, str] = {}
    try:
        with zipfile.ZipFile(archivo_xlsx) as z:
            nombres = set(z.namelist())
            requeridos = {
                "xl/worksheets/sheet1.xml",
                "xl/metadata.xml",
                "xl/richData/richValueRel.xml",
                "xl/richData/_rels/richValueRel.xml.rels",
            }
            if not requeridos.issubset(nombres):
                return {}
            sheet = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
            celda_a_vm: dict[int, int] = {}
            for ref, vm in re.findall(r'<c[^>]*r="([A-Z]+\d+)"[^>]*vm="(\d+)"', sheet):
                if ref.startswith("A") and ref[1:].isdigit():
                    celda_a_vm[int(ref[1:])] = int(vm)
            if not celda_a_vm:
                return {}
            meta_xml = z.read("xl/metadata.xml").decode("utf-8")
            ns_meta = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            root = ET.fromstring(meta_xml)
            value_meta = root.find("x:valueMetadata", ns_meta)
            vm_a_rvb: dict[int, int] = {}
            if value_meta is not None:
                for idx_vm, bk in enumerate(value_meta.findall("x:bk", ns_meta), start=1):
                    rc = bk.find("x:rc", ns_meta)
                    if rc is not None:
                        v = rc.get("v")
                        if v is not None:
                            vm_a_rvb[idx_vm] = int(v)
            rvb_xml = z.read("xl/richData/richValueRel.xml").decode("utf-8")
            ns_rvb = {
                "rvr": "http://schemas.microsoft.com/office/spreadsheetml/2022/richvaluerel",
                "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
            }
            rvb_root = ET.fromstring(rvb_xml)
            rvb_a_rid: dict[int, str] = {}
            for idx, rel in enumerate(rvb_root.findall("rvr:rel", ns_rvb)):
                rid = rel.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
                if rid:
                    rvb_a_rid[idx] = rid
            rels_xml = z.read("xl/richData/_rels/richValueRel.xml.rels").decode("utf-8")
            ns_rels = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
            rels_root = ET.fromstring(rels_xml)
            rid_a_target: dict[str, str] = {}
            for rel in rels_root.findall("r:Relationship", ns_rels):
                rid = rel.get("Id")
                target = rel.get("Target")
                if rid and target:
                    rid_a_target[rid] = target
            for fila, vm in celda_a_vm.items():
                rvb = vm_a_rvb.get(vm)
                if rvb is None:
                    continue
                rid = rvb_a_rid.get(rvb)
                if rid is None:
                    continue
                target = rid_a_target.get(rid)
                if target is None:
                    continue
                if target.startswith("/"):
                    ruta_zip = target.lstrip("/")
                else:
                    ruta_zip = os.path.normpath(
                        os.path.join("xl/richData", target)
                    ).replace("\\", "/")
                if ruta_zip not in nombres:
                    continue
                ext = os.path.splitext(ruta_zip)[1] or ".png"
                ruta_local = os.path.join(tmpdir, f"incell_fila{fila}{ext}")
                with z.open(ruta_zip) as src, open(ruta_local, "wb") as dst:
                    dst.write(src.read())
                resultado[fila] = ruta_local
    except Exception as e:
        print(f"  Aviso: no se pudieron extraer imagenes in-cell ({e})")
        return {}
    return resultado


# --------------------------------------------------------------------------
# Conversion xlsb -> xlsx via LibreOffice headless.
# --------------------------------------------------------------------------

class LibreOfficeNoDisponible(RuntimeError):
    """No hay binario de LibreOffice (soffice) para convertir el .xlsb."""


def _soffice_bin() -> str | None:
    for nombre in ("soffice", "libreoffice"):
        ruta = shutil.which(nombre)
        if ruta:
            return ruta
    return None


def convertir_a_xlsx(src: Path, dest_dir: Path | None = None) -> Path:
    """Convierte un .xlsb/.xls a .xlsx con LibreOffice headless.

    Devuelve la ruta del .xlsx generado. Lanza LibreOfficeNoDisponible si no
    hay soffice, o RuntimeError si la conversion falla.
    """
    src = Path(src)
    dest_dir = Path(dest_dir) if dest_dir else src.parent
    soffice = _soffice_bin()
    if not soffice:
        raise LibreOfficeNoDisponible(
            "No se encontro LibreOffice (soffice) para convertir el template a .xlsx."
        )
    dest_dir.mkdir(parents=True, exist_ok=True)
    # UserInstallation propio y aislado: evita el error "User installation could
    # not be completed" cuando HOME esta vacio o hay instancias concurrentes.
    with tempfile.TemporaryDirectory() as perfil:
        env = {**os.environ, "HOME": perfil}
        proc = subprocess.run(
            [soffice, "--headless", "--norestore", "--nologo", "--nofirststartwizard",
             f"-env:UserInstallation=file://{perfil}/lo_profile",
             "--convert-to", "xlsx", "--outdir", str(dest_dir), str(src)],
            capture_output=True, text=True, env=env, stdin=subprocess.DEVNULL,
        )
    salida = dest_dir / (src.stem + ".xlsx")
    if proc.returncode != 0 or not salida.exists():
        raise RuntimeError(
            f"LibreOffice no pudo convertir {src.name}: {proc.stderr[:300] or proc.stdout[:300]}"
        )
    return salida


def asegurar_template_xlsx(template: Path) -> Path:
    """Devuelve un template .xlsx. Si es .xlsb/.xls, lo convierte (y cachea).

    El .xlsx cacheado se regenera si el original es mas nuevo.
    """
    template = Path(template)
    if template.suffix.lower() == ".xlsx":
        return template
    cache = template.with_suffix(".xlsx")
    if cache.exists() and cache.stat().st_mtime >= template.stat().st_mtime:
        return cache
    return convertir_a_xlsx(template, cache.parent)


# --------------------------------------------------------------------------
# Llenado del formato con openpyxl.
# --------------------------------------------------------------------------

def _detectar_n_productos(ws, filas_con_imagen: set[int], cols_relevantes: list[int]) -> int:
    """Cuenta filas con datos desde la fila 2, parando en la primera vacia."""
    n = 0
    fila = 2
    while fila < 10000:
        tiene_foto = fila in filas_con_imagen
        tiene_valor = any(
            ws.cell(row=fila, column=c).value not in (None, "")
            for c in cols_relevantes if c != 1
        )
        if not tiene_foto and not tiene_valor:
            break
        n += 1
        fila += 1
    return n


def _insertar_imagen_centrada(ws, ruta_png: str, col_idx0: int, fila_idx0: int) -> None:
    """Inserta una imagen escalada y centrada en la celda (col_idx0, fila_idx0),
    ambos 0-indexados. Mantiene el aspecto original."""
    img = XLImage(ruta_png)
    orig_w, orig_h = img.width, img.height
    # Caja objetivo (px) a partir de la geometria de la fila de foto.
    cell_w_px = int(ANCHO_COL_CHARS * 7 + 5)
    cell_h_px = int(ALTO_FILA_FOTO_PT * 96 / 72)
    margen = 6
    box_w = max(1, cell_w_px - 2 * margen)
    box_h = max(1, cell_h_px - 2 * margen)
    ratio = min(box_w / orig_w, box_h / orig_h, 1.0)
    w = max(1, int(orig_w * ratio))
    h = max(1, int(orig_h * ratio))
    off_x = max(margen, (cell_w_px - w) // 2)
    off_y = max(margen, (cell_h_px - h) // 2)
    img.width, img.height = w, h
    marker = AnchorMarker(col=col_idx0, colOff=pixels_to_EMU(off_x),
                          row=fila_idx0, rowOff=pixels_to_EMU(off_y))
    img.anchor = OneCellAnchor(
        _from=marker,
        ext=XDRPositiveSize2D(pixels_to_EMU(w), pixels_to_EMU(h)),
    )
    ws.add_image(img)


def llenar_formato_hd(
    archivo_origen: str,
    template_xlsx: str,
    archivo_salida: str,
    mapeo_datos: dict[int, int],
    constantes: dict[int, str] | None = None,
) -> int:
    """Rellena el template HD (xlsx) con los productos del xlsx intermedio.

    Escribe una columna por producto (C, D, E, ...): CONSTANTES + datos mapeados
    + MPT-XXXX (fila 9) + foto (fila 2). Devuelve el numero de productos escritos.

    `constantes` permite sobreescribir CONSTANTES (ej. Vendor Name/Number por
    proyecto). Si es None se usa el CONSTANTES del modulo.
    """
    consts = CONSTANTES if constantes is None else constantes
    with tempfile.TemporaryDirectory() as tmpdir:
        in_cell = extraer_imagenes_in_cell(archivo_origen, tmpdir)
        flotantes = extraer_imagenes_drawing(archivo_origen, tmpdir)
        fotos: dict[int, str] = {**flotantes, **in_cell}

        wb_o = openpyxl.load_workbook(archivo_origen)
        ws_o = wb_o.active
        cols_relevantes = [1] + list(mapeo_datos.keys())
        n = _detectar_n_productos(ws_o, set(fotos), cols_relevantes)
        if n == 0:
            raise ValueError("El archivo origen no tiene productos detectables (fila 2+).")

        wb_d = openpyxl.load_workbook(template_xlsx)
        ws_d = wb_d.active
        # Quitar las imagenes de muestra del template (fotos de producto de ejemplo).
        ws_d._images = []

        col_final = 3 + n - 1
        # Limpiar el area de productos (cols C..col_final, filas 2-18) para que
        # no queden datos de la columna de muestra si un producto trae vacios.
        for c in range(3, col_final + 1):
            for r in range(2, 19):
                ws_d.cell(row=r, column=c).value = None

        for i in range(n):
            fila_origen = 2 + i
            col_dest = 3 + i
            letra = openpyxl.utils.get_column_letter(col_dest)
            ws_d.column_dimensions[letra].width = ANCHO_COL_CHARS

            for fila_dest, valor in consts.items():
                ws_d.cell(row=fila_dest, column=col_dest, value=valor)
            ws_d.cell(row=9, column=col_dest, value=f"MPT-{i + 1:04d}")
            for col_origen, fila_dest in mapeo_datos.items():
                valor = ws_o.cell(row=fila_origen, column=col_origen).value
                if valor in (None, ""):
                    continue
                ws_d.cell(row=fila_dest, column=col_dest, value=valor)

            for fila_dest in FMT_MONEDA_FILAS:
                ws_d.cell(row=fila_dest, column=col_dest).number_format = FMT_MONEDA
            for fila_dest in FMT_PORCENTAJE_FILAS:
                ws_d.cell(row=fila_dest, column=col_dest).number_format = FMT_PORCENTAJE

            ruta_png = fotos.get(fila_origen)
            if ruta_png and os.path.exists(ruta_png):
                _insertar_imagen_centrada(ws_d, ruta_png, col_dest - 1, FOTO_FILA - 1)

        ws_d.row_dimensions[FOTO_FILA].height = ALTO_FILA_FOTO_PT
        wb_d.save(archivo_salida)
        return n
