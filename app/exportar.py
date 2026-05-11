"""
Genera el xlsx intermedio a partir de los productos marcados en la BD.
Despues, ese xlsx se puede pasar a llenar_formato_hd.py para producir
el formato HD final.

El intermedio tiene 17 columnas:
    A  Foto
    B  SKU
    C  Descripcion
    D  Medidas
    E  Material
    F  Peso (kg)
    G  Color
    H  MOQ
    I  Packing
    J  Carton dims
    K  CBM
    L  Pzas 20ft
    M  Pzas 40hq
    N  Lead time
    O  Venta HD MXN (paso 11 redondeado al entero) -> "DOMESTIC COST" del HD
    P  Retail c/IVA MXN (paso 13 redondeado al entero) -> "SUGGESTED RETAIL"
    Q  Margen HD (decimal 0.30 = 30%, formula 1-venta/(retail/1.16)) -> "THD MARGIN"

Si el producto tiene snapshots guardados, _cotizar_producto usa el mas
reciente para venta y retail (refleja edicion manual del usuario en el
panel) en lugar de los pasos 11/13 default del motor.
"""

from pathlib import Path
import re

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font

from sqlalchemy.orm import Session

from app.modelos import Producto, CostoAdicional, Foto, CotizacionSnapshot


def _resolver_foto_path(producto: Producto, base_fotos: str, db: Session) -> Path | None:
    """Devuelve la ruta del archivo de foto a usar. Si el producto no tiene
    foto propia, intenta usar la del SKU base (sin sufijos tipo -7CM, -L10CM).

    Esto soluciona el caso de variantes clonadas que se quedaron sin fotos.
    """
    base_path = Path(base_fotos)
    # Foto propia primero
    for f in producto.fotos:
        ruta = base_path / f.ruta_relativa
        if ruta.exists():
            return ruta
    # Fallback: buscar otro producto con el mismo SKU base
    if producto.sku:
        # Extraer base: 'PLB002-10CM' -> 'PLB002'; 'PLB002' -> 'PLB002'
        m = re.match(r"^([A-Z0-9]+?)(?:[-_].*)?$", producto.sku.strip().upper())
        sku_base = m.group(1) if m else producto.sku.strip().upper()
        if sku_base and sku_base != producto.sku.upper():
            candidatos = (
                db.query(Producto)
                .filter(Producto.sku == sku_base)
                .all()
            )
            for c in candidatos:
                for f in c.fotos:
                    ruta = base_path / f.ruta_relativa
                    if ruta.exists():
                        return ruta
    return None


IVA_MX = 0.16  # IVA Mexico para invertir retail c/IVA -> sin IVA


def _margen_hd(retail_civa: float, venta_hd: float) -> float:
    """Margen HD = 1 - venta_hd / (retail / 1.16). Devuelve decimal (0.30 = 30%)."""
    if retail_civa <= 0 or venta_hd <= 0:
        return 0.0
    retail_siva = retail_civa / (1 + IVA_MX)
    if retail_siva <= 0:
        return 0.0
    return 1 - venta_hd / retail_siva


def _cotizar_producto(
    producto: Producto,
    fob_efectivo: float,
    db: Session,
    *,
    params: dict | None = None,
) -> dict:
    """Devuelve los datos de cotizacion para exportar.

    Resolucion de fuente:
      1. Si llegan `params` no vacios (usuario movio algo en la barra),
         se IGNORA el snapshot y se recalcula todo con esos params.
         Esto evita que un snapshot viejo congele retail/margenes cuando
         el usuario cambio el TC o los margenes en la barra superior.
      2. Si NO llegan params (request sin query string) y hay snapshot,
         se usa el snapshot (retail editado, venta y margen guardados).
      3. Si no hay nada, motor con defaults.

    Margen HD en los exports SIEMPRE se recalcula con la formula:
        margen_hd = 1 - venta_hd / (retail_civa / 1.16)
    asi refleja el margen real del retailer aun cuando el retail haya sido
    movido manualmente.

    Args:
        params: dict opcional con keys tc, margen_nuestro_pct,
            margen_cliente_pct, flete_maritimo_usd, flete_local_mxn,
            descuentos_pct, descuentos_na_pct, gasto_fijo_pct,
            gastos_aduanales_pct. Cualquier subset.
    """
    params = params or {}
    # Si llegan params no vacios, el snapshot se ignora completamente
    usar_snapshot_si_existe = not bool(params)
    base = {
        "fob_efectivo_usd": fob_efectivo,
        "piezas_contenedor": 0,
        "tasa_arancelaria_pct": 0.0,
        "fraccion": "",
        "tipo_cambio": 0.0,
        "landed_unit_mxn": 0.0,
        "venta_hd_mxn": 0.0,
        "retail_civa_mxn": 0.0,
        "retail_redondeado_mxn": 0.0,
        "margen_lloyds_real": 0.0,
        "margen_cliente_pct": 0.0,
        "fuente": "motor",  # 'motor' | 'snapshot'
    }
    try:
        from app.cotizador.adapter import producto_a_row
        from app.cotizador.engine import compute_for_row
        from app.cotizador.lookup import resolver_arancel
        from app.cotizador.defaults import country_params

        # Buscar snapshot mas reciente
        snap = (
            db.query(CotizacionSnapshot)
            .filter_by(producto_id=producto.id)
            .order_by(CotizacionSnapshot.creado_en.desc())
            .first()
        )

        # Si llegan params explicitos, ignorar snapshot. Si no, usar snapshot
        # como fuente preferida (con fallback a None para usar defaults).
        if usar_snapshot_si_existe and snap:
            snap_params = {
                "tc": snap.tc,
                "margen_nuestro_pct": snap.margen_nuestro_pct,
                "margen_cliente_pct": snap.margen_cliente_pct,
                "flete_maritimo_usd": snap.flete_maritimo_usd,
                "flete_local_mxn": snap.flete_local_mxn,
                "descuentos_pct": snap.descuentos_pct,
                "descuentos_na_pct": snap.descuentos_na_pct,
                "gasto_fijo_pct": snap.gasto_fijo_pct,
                "gastos_aduanales_pct": snap.gastos_aduanales_pct,
            }
        else:
            snap_params = {}

        # Resolver cada param: snapshot > params barra > None (engine usa default)
        def pick(k: str):
            v = snap_params.get(k)
            if v is not None:
                return v
            return params.get(k)

        settings: dict = {}
        if pick("flete_local_mxn") is not None:
            settings["flete_local_mxn"] = pick("flete_local_mxn")
        else:
            settings["flete_local_mxn"] = 70000
        if pick("descuentos_pct") is not None:
            settings["descuentos_pct"] = pick("descuentos_pct")
        if pick("descuentos_na_pct") is not None:
            settings["descuentos_na_pct"] = pick("descuentos_na_pct")
        if pick("gasto_fijo_pct") is not None:
            settings["gasto_fijo_pct"] = pick("gasto_fijo_pct")
        if pick("gastos_aduanales_pct") is not None:
            settings["gastos_aduanales_pct"] = pick("gastos_aduanales_pct")

        row = producto_a_row(producto)
        row["unit_price"] = fob_efectivo
        arancel = resolver_arancel(db, producto.categoria, producto.subcategoria, producto.material)

        res = compute_for_row(
            row,
            settings=settings,
            override_tasa_pct=arancel.tasa_pct,
            override_tc=pick("tc"),
            override_flete_maritimo_usd=pick("flete_maritimo_usd"),
            margen_nuestro_pct=pick("margen_nuestro_pct"),
            margen_cliente_pct=pick("margen_cliente_pct"),
        )

        landed = float(res.paso9)
        venta_motor = float(res.paso11)
        retail_motor = float(res.paso13)
        retail_redondeado = float(res.paso14)

        base.update({
            "piezas_contenedor": int(row["piezas_contenedor"] or 0),
            "tasa_arancelaria_pct": float(arancel.tasa_pct),
            "fraccion": arancel.fraccion or "",
            "tipo_cambio": float(res.tipo_cambio),
            "landed_unit_mxn": landed,
            "retail_redondeado_mxn": retail_redondeado,
        })

        if usar_snapshot_si_existe and snap and snap.retail_final_mxn and snap.retail_final_mxn > 0:
            # Sin params: retail editado del snapshot manda
            retail_civa = float(snap.retail_final_mxn)
            venta_hd = float(snap.venta_lloyds_mxn or venta_motor)
            margen_lloyds = float(snap.margen_real_pct or 0) / 100
            base["fuente"] = "snapshot"
        else:
            # Con params explicitos, o sin snapshot: usar motor con params/defaults
            retail_civa = retail_motor
            venta_hd = venta_motor
            cp = country_params(res.country_code, settings=settings)
            td = (float(cp["descuentos_pct"]) + float(cp["descuentos_na_pct"]) + float(cp["gasto_fijo_pct"])) / 100
            margen_lloyds = (1 - landed / venta_hd - td) if venta_hd > 0 else 0

        base.update({
            "venta_hd_mxn": venta_hd,
            "retail_civa_mxn": retail_civa,
            "margen_lloyds_real": margen_lloyds,
            # Margen HD calculado con la formula pedida:
            # 1 - venta_hd / (retail_civa / 1.16)
            "margen_cliente_pct": _margen_hd(retail_civa, venta_hd) * 100,
        })
    except Exception:
        pass
    return base


def _construir_xlsx_intermedio(
    productos: list,
    xlsx_intermedio: str,
    base_fotos: str,
    db: Session,
    params: dict | None = None,
) -> int:
    """Logica compartida que dado una lista de Producto construye el xlsx."""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cotizacion seleccionados"

    headers = [
        "Foto", "SKU", "Descripcion", "Medidas", "Material", "Peso (kg)",
        "Color", "MOQ", "Packing", "Carton dims", "CBM",
        "Pzas 20ft", "Pzas 40hq", "Lead time", "Venta HD MXN",
        "Retail c/IVA MXN", "Margen HD",
    ]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True)
    ws.row_dimensions[1].height = 25

    for i, p in enumerate(productos, start=2):
        ws.row_dimensions[i].height = 90

        # FOB efectivo = fob_usd + suma costos adicionales
        costos = db.query(CostoAdicional).filter_by(producto_id=p.id).all()
        suma_costos = sum(c.monto_usd for c in costos)
        fob_efectivo = (p.fob_usd or 0) + suma_costos

        # Cotizacion via motor 14 pasos (snapshot > params barra > defaults)
        cot = _cotizar_producto(p, fob_efectivo, db, params=params)

        # Costo HD = paso 11 (venta a HD MXN sin IVA), redondeado al entero
        venta_hd = round(cot["venta_hd_mxn"]) if cot["venta_hd_mxn"] > 0 else 0
        # Retail HD = paso 13 (publico c/IVA), redondeado al entero
        retail_hd = round(cot["retail_civa_mxn"]) if cot["retail_civa_mxn"] > 0 else 0

        ws.cell(i, 2, value=p.sku or "")
        ws.cell(i, 3, value=p.descripcion or "")
        ws.cell(i, 4, value=p.medidas or "")
        ws.cell(i, 5, value=p.material or "")
        ws.cell(i, 6, value=p.peso_kg)
        ws.cell(i, 7, value=p.color or "")
        ws.cell(i, 8, value=p.moq or "")
        ws.cell(i, 9, value=p.packing or "")
        ws.cell(i, 10, value=p.carton_dims or "")
        ws.cell(i, 11, value=p.cbm)
        ws.cell(i, 12, value=p.pzas_20ft)
        ws.cell(i, 13, value=p.pzas_40hq)
        ws.cell(i, 14, value=p.lead_time or "")
        ws.cell(i, 15, value=venta_hd)
        ws.cell(i, 16, value=retail_hd)
        # Fila 17 del HD = THD MARGIN = margen del retailer, calculado con
        # la formula 1 - venta_hd / (retail/1.16). En decimal porque la celda
        # tiene formato '0.00%' en llenar_formato_hd.py.
        ws.cell(i, 17, value=cot["margen_cliente_pct"] / 100)

        # Foto: usar foto propia, o fallback al SKU base si la variante no tiene
        foto_path = _resolver_foto_path(p, base_fotos, db)
        if foto_path:
            try:
                img = XLImage(str(foto_path))
                img.width = min(img.width, 120)
                img.height = min(img.height, 120)
                img.anchor = f"A{i}"
                ws.add_image(img)
            except Exception:
                pass

    Path(xlsx_intermedio).parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_intermedio)
    return len(productos)


def generar_formato_hd_desde_marcados(
    session: Session,
    xlsx_intermedio: str,
    base_fotos: str,
    params: dict | None = None,
) -> int:
    """Construye xlsx intermedio con productos marcado_cotizar=True."""
    productos = (
        session.query(Producto)
        .filter(Producto.marcado_cotizar.is_(True))
        .all()
    )
    return _construir_xlsx_intermedio(productos, xlsx_intermedio, base_fotos, session, params=params)


def generar_formato_hd_por_categoria(
    session: Session,
    xlsx_intermedio: str,
    base_fotos: str,
    categoria: str | None,
    params: dict | None = None,
) -> int:
    """Construye xlsx intermedio filtrando por categoria, sin tocar marcas.

    Si categoria es None, exporta los productos sin categoria.
    """
    q = session.query(Producto)
    if categoria is None:
        q = q.filter(Producto.categoria.is_(None))
    else:
        q = q.filter(Producto.categoria == categoria)
    productos = q.all()
    return _construir_xlsx_intermedio(productos, xlsx_intermedio, base_fotos, session, params=params)


# ============================================================
# Export interno: vertical, todas las columnas por producto
# ============================================================

# Columnas del export interno (en orden, A..)
INTERNO_HEADERS = [
    "Foto",
    "SKU",
    "Descripcion",
    "Proveedor",
    "Categoria",
    "Subcategoria",
    "Material",
    "Medidas",
    "Peso (kg)",
    "Color",
    "MOQ",
    "Packing",
    "Carton dims",
    "CBM / caja",
    "Pzas / 20ft",
    "Pzas / 40HQ",
    "Lead time",
    "FOB original USD",
    "Costos adicionales USD",
    "FOB efectivo USD",
    "Fraccion arancelaria",
    "Tasa arancel %",
    "Tipo de cambio",
    "Landing unit MXN",
    "Venta HD MXN (paso 11)",
    "Margen Lloyds real %",
    "Retail c/IVA MXN (paso 13)",
    "Retail redondeado MXN (paso 14)",
    "Margen HD %",
]


def generar_export_interno_marcados(
    session: Session,
    xlsx_salida: str,
    base_fotos: str,
    params: dict | None = None,
) -> int:
    """Genera un xlsx vertical para uso interno con TODAS las columnas por
    producto marcado. Una fila por producto, headers en A1..AC1.

    Incluye foto incrustada, FOB original + costos adicionales + FOB efectivo,
    arancel, TC, landing, venta a HD, retail con y sin redondeo, margen Lloyds
    real y margen HD.
    """
    productos = (
        session.query(Producto)
        .filter(Producto.marcado_cotizar.is_(True))
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cotizacion interna"

    # Headers
    for i, h in enumerate(INTERNO_HEADERS, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True)
    ws.row_dimensions[1].height = 30

    # Ancho columnas (la foto es la mas ancha)
    ws.column_dimensions["A"].width = 18  # Foto
    ws.column_dimensions["B"].width = 18  # SKU
    ws.column_dimensions["C"].width = 50  # Descripcion
    for letra in ["D", "E", "F", "G"]:
        ws.column_dimensions[letra].width = 16
    for letra in ["H", "I", "J", "K", "L", "M"]:
        ws.column_dimensions[letra].width = 14

    for i, p in enumerate(productos, start=2):
        ws.row_dimensions[i].height = 95

        costos = session.query(CostoAdicional).filter_by(producto_id=p.id).all()
        suma_costos = sum(c.monto_usd for c in costos)
        fob_original = p.fob_usd or 0
        fob_efectivo = fob_original + suma_costos
        cot = _cotizar_producto(p, fob_efectivo, session, params=params)

        # Fila i
        ws.cell(i, 2, value=p.sku or "")
        ws.cell(i, 3, value=p.descripcion or "")
        ws.cell(i, 4, value=p.proveedor.nombre if p.proveedor else "")
        ws.cell(i, 5, value=p.categoria or "")
        ws.cell(i, 6, value=p.subcategoria or "")
        ws.cell(i, 7, value=p.material or "")
        ws.cell(i, 8, value=p.medidas or "")
        ws.cell(i, 9, value=p.peso_kg)
        ws.cell(i, 10, value=p.color or "")
        ws.cell(i, 11, value=p.moq or "")
        ws.cell(i, 12, value=p.packing or "")
        ws.cell(i, 13, value=p.carton_dims or "")
        ws.cell(i, 14, value=p.cbm)
        ws.cell(i, 15, value=p.pzas_20ft)
        ws.cell(i, 16, value=p.pzas_40hq)
        ws.cell(i, 17, value=p.lead_time or "")
        ws.cell(i, 18, value=fob_original)
        ws.cell(i, 19, value=suma_costos)
        ws.cell(i, 20, value=fob_efectivo)
        ws.cell(i, 21, value=cot["fraccion"])
        ws.cell(i, 22, value=cot["tasa_arancelaria_pct"])
        ws.cell(i, 23, value=cot["tipo_cambio"])
        ws.cell(i, 24, value=cot["landed_unit_mxn"])
        ws.cell(i, 25, value=round(cot["venta_hd_mxn"]) if cot["venta_hd_mxn"] > 0 else 0)
        ws.cell(i, 26, value=cot["margen_lloyds_real"] * 100)
        ws.cell(i, 27, value=cot["retail_civa_mxn"])
        ws.cell(i, 28, value=round(cot["retail_civa_mxn"]) if cot["retail_civa_mxn"] > 0 else 0)
        ws.cell(i, 29, value=cot["margen_cliente_pct"])

        # Formato moneda y porcentaje
        for col in [18, 19, 20]:
            ws.cell(i, col).number_format = "$#,##0.00"
        for col in [24, 25, 27, 28]:
            ws.cell(i, col).number_format = "$#,##0.00"
        for col in [22, 26, 29]:
            ws.cell(i, col).number_format = "0.00\"%\""

        # Foto incrustada en columna A
        foto_path = _resolver_foto_path(p, base_fotos, session)
        if foto_path:
            try:
                img = XLImage(str(foto_path))
                img.width = min(img.width, 120)
                img.height = min(img.height, 120)
                img.anchor = f"A{i}"
                ws.add_image(img)
            except Exception:
                pass

    # Freeze header
    ws.freeze_panes = "A2"

    Path(xlsx_salida).parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_salida)
    return len(productos)
