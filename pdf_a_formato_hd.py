"""
Procesa un PDF de cotizacion: extrae via Adobe -> construye xlsx intermedio
con la estructura que llenar_formato_hd.py entiende -> genera formato-hd-XXX.xlsx.

Pipeline:
    1. Adobe Extract API -> ZIP con structuredData.json + tables/*.xlsx + figures/*.png
    2. Parsea structuredData.json para mapear cada fila TR[N] a su Item NO,
       descripcion y foto.
    3. Genera un xlsx intermedio "<nombre>_pdf.xlsx" con:
         - col A: foto (insertada como imagen)
         - col C: descripcion
         - col O: precio FOB Ningbo (cuando se detecta)
    4. Llama a llenar_formato_hd.py con ese intermedio.

Uso:
    python pdf_a_formato_hd.py archivo.pdf

Credenciales Adobe: lee ADOBE_CLIENT_ID y ADOBE_CLIENT_SECRET desde .env.
"""

from __future__ import annotations

import json
import os
import re
import sys
import subprocess
import zipfile
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from dotenv import load_dotenv

from adobe.pdfservices.operation.auth.service_principal_credentials import (
    ServicePrincipalCredentials,
)
from adobe.pdfservices.operation.pdf_services import PDFServices
from adobe.pdfservices.operation.pdf_services_media_type import PDFServicesMediaType
from adobe.pdfservices.operation.io.cloud_asset import CloudAsset
from adobe.pdfservices.operation.pdfjobs.jobs.extract_pdf_job import ExtractPDFJob
from adobe.pdfservices.operation.pdfjobs.params.extract_pdf.extract_element_type import (
    ExtractElementType,
)
from adobe.pdfservices.operation.pdfjobs.params.extract_pdf.extract_pdf_params import (
    ExtractPDFParams,
)
from adobe.pdfservices.operation.pdfjobs.params.extract_pdf.extract_renditions_element_type import (
    ExtractRenditionsElementType,
)
from adobe.pdfservices.operation.pdfjobs.params.extract_pdf.table_structure_type import (
    TableStructureType,
)
from adobe.pdfservices.operation.pdfjobs.result.extract_pdf_result import (
    ExtractPDFResult,
)


# Rutas para el paso 4 (llenar formato HD via Excel COM).
# Se sobreescriben con las vars de entorno SCRIPT_LLENAR_HD y FORMATO_HD_PATH
# (definidas en .env). Si las rutas no existen en runtime, el paso 4 se salta
# automaticamente (util en Linux/Docker, donde no hay Excel COM).
DEFAULT_SCRIPT_LLENAR = (
    r"C:\Users\salomon.DC0\Documents\Mascotas-9Mayo\llenar_formato_hd.py"
)
DEFAULT_FORMATO_HD = (
    r"C:\Users\salomon.DC0\Documents\Mascotas-9Mayo\Formato HD-Mascotas.xlsb"
)


def rutas_formato_hd() -> tuple[str, str]:
    load_dotenv()
    script = os.environ.get("SCRIPT_LLENAR_HD") or DEFAULT_SCRIPT_LLENAR
    formato = os.environ.get("FORMATO_HD_PATH") or DEFAULT_FORMATO_HD
    return script, formato

# Limpiar artefactos de openxml cuando Adobe escribe textos
LIMPIAR_RE = re.compile(r"_x000D_|\r|\n+")


def limpiar(texto: str | None) -> str:
    if not texto:
        return ""
    s = LIMPIAR_RE.sub(" ", str(texto))
    return " ".join(s.split())


def cargar_credenciales() -> tuple[str, str]:
    load_dotenv()
    cid = os.environ.get("ADOBE_CLIENT_ID")
    sec = os.environ.get("ADOBE_CLIENT_SECRET")
    if not cid or not sec:
        sys.exit("Faltan ADOBE_CLIENT_ID/ADOBE_CLIENT_SECRET en .env")
    return cid, sec


def extraer_pdf_con_adobe(pdf_path: str, carpeta_salida: str) -> None:
    cid, sec = cargar_credenciales()
    print("  Autenticando con Adobe...")
    credentials = ServicePrincipalCredentials(client_id=cid, client_secret=sec)
    pdf_services = PDFServices(credentials=credentials)

    print("  Subiendo PDF...")
    with open(pdf_path, "rb") as f:
        stream_asset = pdf_services.upload(
            input_stream=f.read(),
            mime_type=PDFServicesMediaType.PDF,
        )

    print("  Procesando (puede tardar 30-60s)...")
    extract_params = ExtractPDFParams(
        elements_to_extract=[ExtractElementType.TEXT, ExtractElementType.TABLES],
        elements_to_extract_renditions=[
            ExtractRenditionsElementType.TABLES,
            ExtractRenditionsElementType.FIGURES,
        ],
        table_structure_type=TableStructureType.XLSX,
    )
    job = ExtractPDFJob(input_asset=stream_asset, extract_pdf_params=extract_params)
    location = pdf_services.submit(job)
    response = pdf_services.get_job_result(location, ExtractPDFResult)
    result_asset: CloudAsset = response.get_result().get_resource()
    stream = pdf_services.get_content(result_asset)

    os.makedirs(carpeta_salida, exist_ok=True)
    zip_path = os.path.join(carpeta_salida, "_resultado.zip")
    with open(zip_path, "wb") as f:
        f.write(stream.get_input_stream())
    with zipfile.ZipFile(zip_path) as z:
        z.extractall(carpeta_salida)
    print(f"  Resultado: {carpeta_salida}")


# Item NO standalone (PB001, PLB002-1, etc) o al INICIO de un texto largo
RE_ITEM_NO_STANDALONE = re.compile(r"^[A-Z]{2,5}\d{2,5}(?:[-]\d+)?$")
RE_ITEM_NO_PREFIJO = re.compile(r"^([A-Z]{2,5}\d{2,5}(?:[-]\d+)?)\s+(.+)$", re.DOTALL)

# Precio con simbolo $ explicito (siempre se acepta).
RE_PRECIO_CON_SIMBOLO = re.compile(r"^(?:USD?\s*)?\$\s*\d+(?:[.,]\d+)?$", re.IGNORECASE)
# Numero plausible como precio (sin simbolo): 0.30 a 999.99
RE_NUMERO_PRECIO = re.compile(r"^\d{1,3}(?:[.,]\d{1,3})?$")

RE_DIMENSION = re.compile(r"^\s*\d+[.,]?\d*\s*[\*xX]\s*\d", re.IGNORECASE)

# Palabras clave para detectar columnas de precio en headers
PALABRAS_PRECIO = (
    "fob", "exw", "price", "usd", "cost", "precio", "$", "ddp", "cif", "fca"
)


def es_header_precio(texto: str) -> bool:
    """True si el texto del header indica que es una columna de precio."""
    t = texto.lower()
    return any(palabra in t for palabra in PALABRAS_PRECIO)


def clasificar_celda(texto: str, columna_precio: bool = False) -> str:
    """
    Clasifica una celda. Si la columna fue identificada como columna de precio
    por su header, aceptamos numeros sin simbolo $ como precios.
    """
    t = texto.strip()
    if not t:
        return "vacio"
    if RE_ITEM_NO_STANDALONE.match(t):
        return "item"
    if RE_PRECIO_CON_SIMBOLO.match(t):
        return "precio"
    if columna_precio and RE_NUMERO_PRECIO.match(t):
        try:
            v = float(t.replace(",", "."))
            if 0.01 <= v <= 9999:
                return "precio"
        except ValueError:
            pass
    if RE_DIMENSION.match(t):
        return "dim"
    if len(t) > 25:
        return "texto_largo"
    return "texto"


def extraer_codigo_de_descripcion(texto: str) -> tuple[str, str]:
    """
    Si la descripcion empieza con un codigo tipo PB001, lo separa.
    Devuelve (codigo, descripcion_sin_codigo). Si no encuentra, ("", texto).
    """
    if not texto:
        return "", ""
    m = RE_ITEM_NO_PREFIJO.match(texto.strip())
    if m:
        return m.group(1), m.group(2).strip()
    return "", texto.strip()


def parsear_precio_a_float(texto: str) -> float | None:
    if not texto:
        return None
    s = texto.replace(",", "").replace("US", "").replace("$", "").strip()
    m = re.search(r"([0-9]+\.?[0-9]*)", s)
    if not m:
        return None
    try:
        return float(m.group(1))
    except ValueError:
        return None


def parsear_filas(json_path: str) -> dict[int, dict]:
    """
    Lee structuredData.json y devuelve {n_fila_TR: {item_no, desc, fob, foto}}.

    Estrategia robusta: en vez de fiarnos de los indices de columna (que se
    desplazan por celdas merged), clasificamos cada celda por contenido:
      - codigo tipo XX###  -> item_no
      - 'US$X.YZ' o numero plausible como precio -> precio
      - texto largo sin patron -> descripcion
    De los precios de la fila tomamos el MENOR como FOB (heuristica: el mas
    barato suele ser el de container completo).
    """
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)

    celdas: dict[int, dict[int, list[str]]] = defaultdict(lambda: defaultdict(list))
    fotos: dict[int, str] = {}
    fila_max = 0

    # Para mapeo por coordenadas (cuando la figura esta fuera de tabla):
    # Para cada fila TR, guardar (pagina, Y_centro) para hacer matching
    fila_coord: dict[int, tuple[int, float]] = {}
    figuras_huerfanas: list[tuple[int, float, str]] = []  # (page, y, ruta_imagen)

    re_td = re.compile(r"/TR\[(\d+)\]/(TD|TH)(?:\[(\d+)\])?")
    re_tr_solo = re.compile(r"/Table/TR\[(\d+)\]")

    for e in data.get("elements", []):
        p = e.get("Path", "")
        bounds = e.get("Bounds")
        page = e.get("Page", 0)

        # Figura fuera de tabla -> guardar para matching por coordenadas
        if "/Figure" in p and "/Table/" not in p:
            files = e.get("filePaths") or []
            if files and bounds:
                y_centro = (bounds[1] + bounds[3]) / 2
                figuras_huerfanas.append((page, y_centro, files[0]))
            continue

        m_tr = re_tr_solo.search(p)
        if not m_tr:
            continue
        n_tr = int(m_tr.group(1))
        fila_max = max(fila_max, n_tr)

        # Guardar coord centro de la fila para matching posterior
        if bounds and n_tr not in fila_coord:
            y_centro = (bounds[1] + bounds[3]) / 2
            fila_coord[n_tr] = (page, y_centro)

        m = re_td.search(p)
        if not m:
            continue
        col = int(m.group(3)) if m.group(3) else 1

        if "/Figure" in p:
            files = e.get("filePaths") or []
            if files and n_tr not in fotos:
                fotos[n_tr] = files[0]
            continue

        texto = limpiar(e.get("Text"))
        if not texto:
            continue
        celdas[n_tr][col].append(texto)

    # Asignar figuras huerfanas a la fila con Y mas cercana en la misma pagina
    if figuras_huerfanas and fila_coord:
        for page_fig, y_fig, ruta in figuras_huerfanas:
            mejor_tr = None
            mejor_dist = float("inf")
            for n_tr, (page_fila, y_fila) in fila_coord.items():
                if page_fila != page_fig:
                    continue
                dist = abs(y_fila - y_fig)
                if dist < mejor_dist:
                    mejor_dist = dist
                    mejor_tr = n_tr
            # Solo aceptar si la distancia es razonable (< 100 puntos = ~3.5cm)
            if mejor_tr is not None and mejor_dist < 100 and mejor_tr not in fotos:
                fotos[mejor_tr] = ruta

    # Detectar columnas de precio a partir de la fila header
    # Header = la fila TR mas pequena con >= 3 columnas con texto y todos los
    # textos cortos (max ~30 chars).
    columnas_precio: set[int] = set()
    header_info: dict[int, str] = {}
    for n_tr in sorted(celdas.keys()):
        cols = celdas[n_tr]
        if len(cols) < 3:
            continue
        # Validar que los textos sean header-like (todos cortos)
        textos_por_col = {c: " ".join(t).strip() for c, t in cols.items()}
        if all(len(v) <= 50 for v in textos_por_col.values()):
            for col, texto in textos_por_col.items():
                header_info[col] = texto
                if es_header_precio(texto):
                    columnas_precio.add(col)
            break  # solo el primer header valido

    if columnas_precio:
        cols_str = ", ".join(
            f"{c}({header_info.get(c,'?')[:20]})" for c in sorted(columnas_precio)
        )
        print(f"  Columnas precio detectadas: {cols_str}")

    resultado: dict[int, dict] = {}
    for n_tr in range(1, fila_max + 1):
        td = celdas.get(n_tr, {})
        if not td and n_tr not in fotos:
            continue

        # Reunir todas las celdas (col, texto) en orden, clasificando con
        # conocimiento de si la columna es de precio (por su header)
        celdas_fila: list[tuple[int, str, str]] = []  # (col, texto, etiqueta)
        for col in sorted(td.keys()):
            texto_celda = " ".join(td[col]).strip()
            if texto_celda:
                etiq = clasificar_celda(texto_celda, columna_precio=(col in columnas_precio))
                celdas_fila.append((col, texto_celda, etiq))

        # Detectar item_no: primer texto que matchee codigo, o prefijo de un texto largo
        item_no = ""
        for col, t, etiq in celdas_fila:
            if etiq == "item":
                item_no = t
                break

        # Descripcion = primer texto_largo
        desc = ""
        for col, t, etiq in celdas_fila:
            if etiq == "texto_largo":
                desc = t
                break
        if not desc:
            partes = [t for col, t, etiq in celdas_fila if etiq == "texto" and t != item_no]
            desc = " ".join(partes[:3])

        # Si no encontre item_no pero la descripcion empieza con codigo, extraerlo
        if not item_no and desc:
            posible_item, desc_limpia = extraer_codigo_de_descripcion(desc)
            if posible_item:
                item_no = posible_item
                desc = desc_limpia

        # Detectar precios: solo celdas con simbolo $ explicito
        precios: list[float] = []
        for col, t, etiq in celdas_fila:
            if etiq == "precio":
                v = parsear_precio_a_float(t)
                if v is not None:
                    precios.append(v)

        # FOB = el menor (heuristica: FOB para container completo es el barato)
        fob = min(precios) if precios else None

        foto = fotos.get(n_tr)

        if not (item_no or desc or fob is not None or foto):
            continue

        resultado[n_tr] = {
            "item_no": item_no,
            "desc": desc,
            "fob": fob,
            "fob_str": f"{fob}" if fob is not None else "",
            "foto": foto,
            "todo": {col: " ".join(textos) for col, textos in td.items()},
        }

    # Propagar foto a variantes: si una fila no tiene foto pero la anterior si,
    # hereda. Se reinicia al encontrar un nuevo item_no (nuevo producto).
    foto_actual = None
    item_actual = ""
    propagadas = 0
    for n_tr in sorted(resultado.keys()):
        f = resultado[n_tr]
        if f["item_no"] and f["item_no"] != item_actual:
            # Nuevo producto: la foto se reinicia (la tomara este o sus variantes)
            item_actual = f["item_no"]
            foto_actual = f["foto"]
        else:
            # Variante o continuacion: si tiene foto propia, actualiza; si no, hereda
            if f["foto"]:
                foto_actual = f["foto"]
            elif foto_actual:
                f["foto"] = foto_actual
                f["foto_heredada"] = True
                propagadas += 1

    print(f"  Filas con contenido: {len(resultado)} (max TR={fila_max})")
    print(f"  Filas con foto propia: {sum(1 for r in resultado.values() if r['foto'] and not r.get('foto_heredada'))}")
    print(f"  Filas con foto heredada: {propagadas}")
    print(f"  Filas con foto (total): {sum(1 for r in resultado.values() if r['foto'])}")
    print(f"  Filas con item_no: {sum(1 for r in resultado.values() if r['item_no'])}")
    print(f"  Filas con FOB: {sum(1 for r in resultado.values() if r['fob'] is not None)}")
    return resultado




def construir_xlsx_intermedio(
    filas: dict[int, dict],
    carpeta_extract: str,
    xlsx_out: str,
) -> int:
    """
    Construye un xlsx con la misma estructura que llenar_formato_hd.py espera:
      - Hoja 1, fila 1 con encabezados, datos desde fila 2.
      - Col A = foto (imagen incrustada)
      - Col C = descripcion
      - Col O = precio FOB (numero)
      - Col B = Item NO (informativo)
    Devuelve la cantidad de filas escritas.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cotizacion PDF"

    # Headers
    headers = {1: "Foto", 2: "Item NO", 3: "Descripcion"}
    for c in range(4, 15):
        headers[c] = ""
    headers[15] = "FOB"  # col O
    for c, h in headers.items():
        ws.cell(row=1, column=c, value=h)

    # Ajustar alto/ancho razonable
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["O"].width = 12

    fila_excel = 2
    contador = 0
    for n_tr in sorted(filas.keys()):
        f = filas[n_tr]
        # Saltar headers / sub-headers de la tabla del PDF
        item = f["item_no"]
        desc = f["desc"]
        # Filtrar headers del PDF: filas sin item_no parecido a codigo (con letras+numeros)
        if not (f["foto"] or item or desc):
            continue

        ws.row_dimensions[fila_excel].height = 90  # da espacio para la foto

        ws.cell(row=fila_excel, column=2, value=item)
        ws.cell(row=fila_excel, column=3, value=desc)

        if f["fob"] is not None:
            ws.cell(row=fila_excel, column=15, value=f["fob"])

        # Insertar foto si existe
        if f["foto"]:
            ruta_foto = os.path.join(carpeta_extract, f["foto"])
            if os.path.exists(ruta_foto):
                try:
                    img = XLImage(ruta_foto)
                    # Limitar tamano para que no exploda
                    img.width = min(img.width, 120)
                    img.height = min(img.height, 120)
                    img.anchor = f"A{fila_excel}"
                    ws.add_image(img)
                except Exception as e:
                    print(f"    Aviso: no pude insertar foto fila {fila_excel}: {e}")

        fila_excel += 1
        contador += 1

    wb.save(xlsx_out)
    return contador


def construir_xlsx_desde_claude(
    productos: list[dict],
    carpeta_extract: str,
    xlsx_out: str,
) -> int:
    """
    Construye xlsx intermedio a partir de la lista de productos extraida por Claude.
    Mismo layout que construir_xlsx_intermedio (col A foto, B item, C desc, O FOB).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cotizacion PDF"

    # Layout:
    #   A=Foto, B=SKU, C=Descripcion (usado por llenar_formato), D=Medidas,
    #   E=Material, F=Peso(kg), G=Color, H=MOQ, I=Packing, J=Carton size,
    #   K=CBM, L=pzas 20ft, M=pzas 40hq, N=Lead time, O=FOB USD (usado por llenar_formato),
    #   P=Pzas/caja (al final para no desplazar columnas legacy).
    headers = {
        1: "Foto", 2: "SKU", 3: "Descripcion", 4: "Medidas",
        5: "Material", 6: "Peso (kg)", 7: "Color", 8: "MOQ",
        9: "Packing", 10: "Carton dims", 11: "CBM",
        12: "Pzas 20ft", 13: "Pzas 40hq", 14: "Lead time",
        15: "FOB USD", 16: "Pzas/caja",
    }
    for c, h in headers.items():
        ws.cell(row=1, column=c, value=h)

    # Anchos para que se lea bien
    anchos = {"A": 18, "B": 14, "C": 50, "D": 22, "E": 15, "F": 10,
              "G": 25, "H": 14, "I": 22, "J": 22, "K": 10,
              "L": 10, "M": 10, "N": 14, "O": 12, "P": 10}
    for col, w in anchos.items():
        ws.column_dimensions[col].width = w

    # Header row bold
    from openpyxl.styles import Font, Alignment
    for c in headers:
        ws.cell(row=1, column=c).font = Font(bold=True)
        ws.cell(row=1, column=c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    # Helper tolerante a tipos: None -> "", str -> str.strip(),
    # numero (int/float) -> numero. Necesario porque el flow JSON-based de
    # Claude devuelve strings, pero Claude Vision devuelve numeros JSON.
    def _val(v):
        if v is None:
            return ""
        if isinstance(v, str):
            return v.strip()
        return v

    fila_excel = 2
    contador = 0
    for p in productos:
        sku = _val(p.get("sku")) or ""
        if not isinstance(sku, str):
            sku = str(sku)
        desc = _val(p.get("desc")) or ""
        if not isinstance(desc, str):
            desc = str(desc)
        fob = p.get("fob")
        foto_rel = _val(p.get("foto")) or ""
        if not isinstance(foto_rel, str):
            foto_rel = str(foto_rel)

        if not (sku or desc or fob is not None or foto_rel):
            continue

        ws.row_dimensions[fila_excel].height = 90
        ws.cell(row=fila_excel, column=2, value=sku)
        ws.cell(row=fila_excel, column=3, value=desc)
        ws.cell(row=fila_excel, column=4, value=_val(p.get("medidas")))
        ws.cell(row=fila_excel, column=5, value=_val(p.get("material")))
        ws.cell(row=fila_excel, column=6, value=_val(p.get("peso_kg")))
        ws.cell(row=fila_excel, column=7, value=_val(p.get("color")))
        ws.cell(row=fila_excel, column=8, value=_val(p.get("moq")))
        ws.cell(row=fila_excel, column=9, value=_val(p.get("packing")))
        ws.cell(row=fila_excel, column=10, value=_val(p.get("carton")))
        ws.cell(row=fila_excel, column=11, value=_val(p.get("cbm")))
        ws.cell(row=fila_excel, column=12, value=_val(p.get("pzas_20ft")))
        ws.cell(row=fila_excel, column=13, value=_val(p.get("pzas_40hq")))
        ws.cell(row=fila_excel, column=14, value=_val(p.get("lead_time")))
        if fob is not None:
            try:
                ws.cell(row=fila_excel, column=15, value=float(fob))
            except (TypeError, ValueError):
                pass
        ws.cell(row=fila_excel, column=16, value=_val(p.get("pzas_caja")))

        # Wrap text en celdas de texto
        for col in (3, 4, 7, 9, 10):
            ws.cell(row=fila_excel, column=col).alignment = Alignment(
                wrap_text=True, vertical="center"
            )

        if foto_rel:
            ruta_foto = os.path.join(carpeta_extract, foto_rel)
            if os.path.exists(ruta_foto):
                try:
                    img = XLImage(ruta_foto)
                    img.width = min(img.width, 120)
                    img.height = min(img.height, 120)
                    img.anchor = f"A{fila_excel}"
                    ws.add_image(img)
                except Exception as e:
                    print(f"    Aviso: no pude insertar foto fila {fila_excel}: {e}")

        fila_excel += 1
        contador += 1

    wb.save(xlsx_out)
    return contador


def calidad_extraccion(filas: dict) -> float:
    """
    Devuelve 0.0-1.0 segun cuantas filas tienen datos utilizables.

    Pesos:
    - 50% pondera fila con FOB (precio = lo mas importante)
    - 30% pondera fila con item_no (identificacion del producto)
    - 20% pondera fila con foto propia (no heredada)

    Si calidad < 0.40 conviene usar fallback con Claude.
    """
    if not filas:
        return 0.0
    n = len(filas)
    con_fob = sum(1 for r in filas.values() if r.get("fob") is not None)
    con_item = sum(1 for r in filas.values() if r.get("item_no"))
    con_foto_propia = sum(
        1 for r in filas.values()
        if r.get("foto") and not r.get("foto_heredada")
    )
    score = (con_fob / n) * 0.5 + (con_item / n) * 0.3 + (con_foto_propia / n) * 0.2
    return score


def main() -> None:
    if len(sys.argv) < 2:
        sys.exit("Uso: python pdf_a_formato_hd.py archivo.pdf")
    pdf_path = os.path.abspath(sys.argv[1])
    if not os.path.exists(pdf_path):
        sys.exit(f"No existe: {pdf_path}")

    base = Path(pdf_path).stem
    # Limpiar nombre: no chars raros, no __ dobles, max 60 chars
    base_corto = re.sub(r"[^\w\-]+", "_", base)
    base_corto = re.sub(r"_+", "_", base_corto).strip("_")[:60].rstrip("_")
    carpeta_extract = os.path.join(
        Path(pdf_path).parent, f"_adobe_extract_{base_corto}"
    )
    xlsx_intermedio = os.path.join(
        Path(pdf_path).parent, f"_intermedio_{base_corto}.xlsx"
    )

    print(f"PDF origen: {pdf_path}")
    print()

    # Paso 1: extraer con Adobe (si no esta ya extraido)
    json_path = os.path.join(carpeta_extract, "structuredData.json")
    if os.path.exists(json_path):
        print(f"1. Ya existe extraccion previa en {carpeta_extract}, reuso.")
    else:
        print("1. Extrayendo con Adobe...")
        extraer_pdf_con_adobe(pdf_path, carpeta_extract)

    # Paso 2: parsear JSON con parser heuristico
    print()
    print("2. Parseando structuredData.json (parser heuristico)...")
    filas = parsear_filas(json_path)
    calidad = calidad_extraccion(filas)
    print(f"   Calidad: {calidad:.0%}")

    # Decidir si usar Claude como fallback
    usar_claude = "--claude" in sys.argv or (
        "--no-claude" not in sys.argv and calidad < 0.40
    )

    # Paso 3: construir xlsx intermedio
    print()
    # El nombre del PDF original siempre va al meta, sirva o no Claude.
    # Con esto el ingest sabe que PDF muestra al hacer click en "Cotizacion original".
    meta: dict = {"pdf_original": Path(pdf_path).name}
    if usar_claude:
        print("3. Calidad baja: usando Claude Haiku como fallback...")
        try:
            from extraer_con_claude import extraer_con_claude
            figures_dir = os.path.join(carpeta_extract, "figures")
            resultado = extraer_con_claude(json_path, figures_dir)
            productos = resultado["productos"]
            meta.update({
                "seller": resultado.get("seller", ""),
                "buyer": resultado.get("buyer", ""),
            })
            n = construir_xlsx_desde_claude(productos, carpeta_extract, xlsx_intermedio)
            print(f"   {n} productos escritos en {xlsx_intermedio}")
        except Exception as e:
            print(f"   Claude fallo ({e}), volviendo a parser heuristico")
            n = construir_xlsx_intermedio(filas, carpeta_extract, xlsx_intermedio)
            print(f"   {n} filas escritas.")
    else:
        print(f"3. Generando xlsx intermedio: {xlsx_intermedio}")
        n = construir_xlsx_intermedio(filas, carpeta_extract, xlsx_intermedio)
        print(f"   {n} filas escritas.")

    # Paso 3.5: Vision fallback. Si despues de parser + Claude/JSON aun
    # quedan 0 productos, intentamos renderizar las paginas como PNG y
    # pedirle a Claude Vision que las lea directamente. Sirve para PDFs
    # sin estructura tabular detectable por Adobe Extract.
    if n == 0 and "--no-vision" not in sys.argv:
        print()
        print("3.5. Sin productos extraidos. Probando Claude Vision como ultimo fallback...")
        try:
            from extraer_con_claude import extraer_con_vision
            resultado_v = extraer_con_vision(pdf_path)
            productos_v = resultado_v.get("productos") or []
            if productos_v:
                # Vision puede haber detectado seller/buyer que JSON no vio
                if resultado_v.get("seller"):
                    meta["seller"] = resultado_v["seller"]
                if resultado_v.get("buyer"):
                    meta["buyer"] = resultado_v["buyer"]
                n = construir_xlsx_desde_claude(productos_v, carpeta_extract, xlsx_intermedio)
                print(f"   {n} productos escritos por Vision en {xlsx_intermedio}")
            else:
                print("   Vision tampoco extrajo productos. Termina con 0 filas.")
        except Exception as e:
            print(f"   Vision fallo: {e}")

    # Persistir metadata (seller/buyer) junto al intermedio.
    # El ingest la leera para nombrar al proveedor.
    meta_path = xlsx_intermedio + ".meta.json"
    if meta:
        Path(meta_path).write_text(
            json.dumps(meta, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(f"   Metadata: {meta_path}")

    # Paso 4: invocar llenar_formato_hd.py
    # Saltar con --no-hd (necesario en Linux/Docker donde no hay Excel COM).
    if "--no-hd" in sys.argv:
        print()
        print(f"4. Saltado (--no-hd). Intermedio listo en: {xlsx_intermedio}")
        sys.exit(0)

    script_llenar, formato_hd = rutas_formato_hd()
    # Auto-skip si las rutas no existen (Linux/Docker o paths cambiaron).
    if not os.path.exists(script_llenar) or not os.path.exists(formato_hd):
        print()
        print("4. Saltado: SCRIPT_LLENAR_HD/FORMATO_HD_PATH no existen en el sistema.")
        print(f"   script: {script_llenar} ({'OK' if os.path.exists(script_llenar) else 'NO EXISTE'})")
        print(f"   formato: {formato_hd} ({'OK' if os.path.exists(formato_hd) else 'NO EXISTE'})")
        print(f"   Intermedio listo en: {xlsx_intermedio}")
        sys.exit(0)

    print()
    print("4. Llamando a llenar_formato_hd.py...")
    cmd = [
        sys.executable,
        script_llenar,
        xlsx_intermedio,
        formato_hd,
        "--mapeo",
        "C=8,O=11",
    ]
    print(f"   {' '.join(cmd)}")
    res = subprocess.run(cmd, capture_output=False)
    sys.exit(res.returncode)


if __name__ == "__main__":
    main()
