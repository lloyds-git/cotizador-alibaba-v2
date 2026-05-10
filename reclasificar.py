#!/usr/bin/env python3
"""Build Mascotas-9Mayo INDICE.xlsx from saved Gmail thread metadata."""
import json
import os
import re
from datetime import datetime
from pathlib import Path

WORK_DIR = Path(__file__).parent
PAGES = ["threads_page1.json", "threads_page2.json", "threads_page3.json"]

CATEGORIES = [
    "alimentadores",
    "rejas",
    "pajaros",
    "viaje",
    "correas",
    "higiene",
    "camas",
    "juguetes",
    "alimentacion",
]

# Keyword patterns per category (lowercase). Include EN/ES + variations.
KEYWORDS = {
    "alimentadores": [
        r"feeder", r"dispenser", r"fountain", r"bebedero", r"\bbowl\b", r"\bplato",
        r"comedor", r"comedero", r"\belevated\b.*\bbowl\b", r"water dispenser",
        r"\btazon", r"\bdrink\b", r"\btaz[óo]n", r"\bbowls?\b",
        r"silicone.*bowl", r"feeding bowl", r"\bbebed",
    ],
    "rejas": [
        r"\bgate\b", r"\bfence\b", r"\breja", r"pet door", r"\bdoor\b", r"\bjaula\b",
        r"\bpuerta", r"safety gate", r"metal gate",
    ],
    "pajaros": [
        r"\bbird\b", r"hummingbird", r"colibr[íi]", r"\bp[áa]jaro", r"bird feeder",
        r"bird nest", r"birds?",
    ],
    "viaje": [
        r"backpack", r"\bmochila", r"transportador", r"\bcarrier\b", r"car seat",
        r"silla coche", r"pet bag", r"\btravel\b", r"\basiento", r"\bback pack\b",
        r"car nest", r"\bdog bag\b", r"caja transport", r"\bsilla\b", r"\bcaja\b",
        r"\bsilla.*coche", r"pet shoes",
    ],
    "correas": [
        r"\bleash\b", r"\bcorrea", r"\bharness\b", r"\bcollar", r"\barn[ée]s\b",
        r"chest and back", r"pecho y espalda",
    ],
    "higiene": [
        r"shampoo", r"\bsoap\b", r"\btoilet\b", r"\bpotty\b", r"\bba[ñn]o\b",
        r"cleaning", r"\bwipes\b", r"\bpads\b", r"dog potty", r"dispensador de jab[óo]n",
        r"soap dispenser",
    ],
    "camas": [
        r"\bbed\b", r"\bcama", r"cooling bed", r"\bmat\b", r"\btapete", r"\bcuddle\b",
        r"pet bed", r"\bcooling\b", r"silicone pet mat", r"pet mat", r"asiento grande",
        r"\bnido\b", r"foldable.*bed",
    ],
    "juguetes": [
        r"\btoy\b", r"\bjuguete", r"\bplush\b", r"\bpeluche", r"\bball\b",
        r"\btoys\b",
    ],
    "alimentacion": [
        r"food storage", r"food container", r"\bpet food\b", r"\bcomida\b", r"\bpienso\b",
        r"\btreats?\b", r"rice container", r"food box", r"food storage box",
        r"\bdog food\b", r"\bcat food\b", r"food container storage",
        r"platos de comida", r"tapete comida",
    ],
}

# Extra patterns for new (custom) categories
EXTRA_CATEGORIES = {
    "pasto/cesped": [r"\bpasto\b", r"\bcesped\b", r"\bgrass\b", r"\bturf\b"],
    "ropa-zapatos": [r"\bshoes\b", r"\bzapatos\b", r"\bclothes\b", r"chest and back"],
    "silicona": [r"silicon[ae]", r"\bsilicona\b"],
    "bolsas": [r"bolsas? pet", r"\bbag\b.*\bpet\b", r"bolsa", r"langlang.*pet bag"],
    "casa-jaula": [r"dog house", r"plastic house", r"pet house", r"\bnest\b"],
    "fuente-agua": [r"fountain", r"fuente", r"water dispenser"],
}

def classify(subject: str, snippet: str) -> tuple[list[str], str]:
    """Return (categories, source_used). Categories list may be empty."""
    text = (subject + " " + snippet).lower()
    cats = []
    for cat, patterns in KEYWORDS.items():
        for p in patterns:
            if re.search(p, text, re.IGNORECASE):
                if cat not in cats:
                    cats.append(cat)
                break
    # Try extras only if no main category matched OR add specific extras
    extras_found = []
    for cat, patterns in EXTRA_CATEGORIES.items():
        for p in patterns:
            if re.search(p, text, re.IGNORECASE):
                if cat not in extras_found:
                    extras_found.append(cat)
                break

    # Refinement rules:
    # If "fountain" or "fuente" → also alimentadores
    if "fuente-agua" in extras_found and "alimentadores" not in cats:
        cats.append("alimentadores")
    # Bag-pet → viaje
    if "bolsas" in extras_found and "viaje" not in cats:
        cats.append("viaje")
    # Pet shoes → ropa-zapatos
    # Casa/jaula often goes together with rejas? No, leave separate
    # Pet bed mat — already in camas
    # Bird nest — pajaros
    if re.search(r"bird nest", text, re.IGNORECASE) and "pajaros" not in cats:
        cats.append("pajaros")
    # Dog bowl, pet bowl → alimentadores (already covered)

    # Combine main + extras (but extras only as suffix if no main hit)
    if not cats and extras_found:
        cats = extras_found
    elif extras_found:
        # Add specific extras like silicona for visibility
        for e in extras_found:
            if e in ("silicona", "casa-jaula", "ropa-zapatos") and e not in cats:
                cats.append(e)

    return cats, "subject+snippet"


def load_threads():
    all_threads = []
    for p in PAGES:
        path = WORK_DIR / p
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        for t in data.get("threads", []):
            all_threads.append(t)
    return all_threads


def build_rows(threads):
    """One row per thread. (Could be one per attachment if we had the FN list.)"""
    rows = []
    for t in threads:
        thread_id = t["id"]
        msgs = t.get("messages", [])
        # First message gives basic info; combine subjects/snippets across thread
        subjects = []
        snippets = []
        date_min = None
        date_max = None
        senders = set()
        for m in msgs:
            subjects.append(m.get("subject", ""))
            snippets.append(m.get("snippet", ""))
            senders.add(m.get("sender", ""))
            d = m.get("date", "")
            try:
                dt = datetime.strptime(d.replace("Z", ""), "%Y-%m-%dT%H:%M:%S")
                if date_min is None or dt < date_min:
                    date_min = dt
                if date_max is None or dt > date_max:
                    date_max = dt
            except Exception:
                pass

        subj_main = subjects[0] if subjects else ""
        snip_combined = " | ".join([s for s in snippets if s])[:500]

        cats, _ = classify(subj_main + " " + " ".join(subjects[1:]), snip_combined)
        cats_str = ", ".join(cats) if cats else "Sin clasificar"

        # Detect attachment likelihood from subject hints (.xlsx/.pdf/.docx in subject)
        likely_attachment = bool(
            re.search(r"\.(pdf|xlsx|xls|docx|doc)\b", subj_main, re.IGNORECASE)
            or any("Fwd:" in s or "Fw:" in s for s in subjects)
            or "quotation" in (subj_main + snip_combined).lower()
            or "cotiza" in (subj_main + snip_combined).lower()
        )

        # Extract supplier/sender from snippet (look for "De: X" or sender)
        supplier = ""
        for s in snippets:
            m = re.search(r"De:\s*([^<\n]+?)<", s)
            if m:
                supplier = m.group(1).strip()
                break
            m = re.search(r"From:\s*([^<\n]+?)<", s)
            if m:
                supplier = m.group(1).strip()
                break

        # Filename hint from subject
        fn_match = re.search(r"([\w\-\s]+\.(?:pdf|xlsx|xls|docx|doc))", subj_main, re.IGNORECASE)
        filename_hint = fn_match.group(1).strip() if fn_match else ""

        gmail_url = f"https://mail.google.com/mail/u/0/#search/from%3Afortuna%40lloydselectronica.com/{thread_id}"

        rows.append({
            "Fecha": (date_max or date_min or "").strftime("%Y-%m-%d %H:%M") if (date_max or date_min) else "",
            "Asunto": subj_main,
            "Categoria(s)": cats_str,
            "Probable adjunto": "Sí" if likely_attachment else "?",
            "Filename hint": filename_hint,
            "Proveedor": supplier,
            "Mensajes en hilo": len(msgs),
            "Snippet": snip_combined,
            "Gmail Link": gmail_url,
            "Thread ID": thread_id,
        })
    return rows


def write_excel(rows, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Indice Mascotas"

    headers = list(rows[0].keys()) if rows else []
    # Header style
    hdr_fill = PatternFill("solid", fgColor="2F5496")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = border

    for ri, row in enumerate(rows, 2):
        for ci, h in enumerate(headers, 1):
            v = row.get(h, "")
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if h == "Gmail Link" and isinstance(v, str) and v.startswith("http"):
                cell.hyperlink = v
                cell.font = Font(color="2F5496", underline="single")
            # Color-code categories
            if h == "Categoria(s)":
                if v == "Sin clasificar":
                    cell.fill = PatternFill("solid", fgColor="FFE699")
                else:
                    cell.fill = PatternFill("solid", fgColor="E2EFDA")

    # Column widths
    widths = {
        "Fecha": 18,
        "Asunto": 60,
        "Categoria(s)": 28,
        "Probable adjunto": 14,
        "Filename hint": 35,
        "Proveedor": 22,
        "Mensajes en hilo": 10,
        "Snippet": 70,
        "Gmail Link": 20,
        "Thread ID": 22,
    }
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(h, 18)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # Auto filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"

    # Stats sheet
    ws2 = wb.create_sheet("Resumen")
    ws2["A1"] = "Resumen por Categoría"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A3"] = "Categoría"
    ws2["B3"] = "# Hilos"
    ws2["A3"].font = ws2["B3"].font = Font(bold=True)

    cat_count = {}
    for r in rows:
        for c in [c.strip() for c in r["Categoria(s)"].split(",")]:
            cat_count[c] = cat_count.get(c, 0) + 1
    for i, (c, n) in enumerate(sorted(cat_count.items(), key=lambda x: -x[1]), 4):
        ws2.cell(row=i, column=1, value=c)
        ws2.cell(row=i, column=2, value=n)

    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 12

    wb.save(out_path)
    return cat_count


def main():
    threads = load_threads()
    print(f"Loaded {len(threads)} threads")
    rows = build_rows(threads)
    out = Path(__file__).parent / "INDICE.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    cat_count = write_excel(rows, out)
    print(f"Wrote {out} with {len(rows)} rows")
    print("Stats:")
    for c, n in sorted(cat_count.items(), key=lambda x: -x[1]):
        print(f"  {c}: {n}")


if __name__ == "__main__":
    main()
