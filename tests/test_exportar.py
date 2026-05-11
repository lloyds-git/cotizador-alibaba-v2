from pathlib import Path
from app.exportar import generar_formato_hd_desde_marcados
from app.modelos import Proveedor, Producto, Foto


def test_exportar_marcados(db_session, tmp_path):
    prov = Proveedor(nombre="V", archivo_pdf="v.pdf")
    db_session.add(prov)
    db_session.commit()

    # pzas_40hq y cbm completos para que el motor 14 pasos pueda calcular
    # venta HD y retail (columnas O y P del intermedio).
    p1 = Producto(
        proveedor_id=prov.id, sku="A1", descripcion="Desc A1",
        fob_usd=10.0, marcado_cotizar=True,
        pzas_40hq=1000, cbm=0.05, peso_kg=1.0,
    )
    p2 = Producto(
        proveedor_id=prov.id, sku="A2", descripcion="Desc A2",
        fob_usd=20.0, marcado_cotizar=False,
    )
    db_session.add_all([p1, p2])
    db_session.commit()

    # Foto fake
    foto_dir = tmp_path / "fotos"
    foto_dir.mkdir()
    png = foto_dir / "p1.png"
    png.write_bytes(bytes.fromhex(
        "89504e470d0a1a0a0000000d49484452000000010000000108060000001f15c4"
        "890000000a49444154789c6300010000000500010d0a2db40000000049454e44ae426082"
    ))
    db_session.add(Foto(producto_id=p1.id, ruta_relativa="fotos/p1.png", es_principal=True))
    db_session.commit()

    out = tmp_path / "_intermedio_marcados.xlsx"
    n = generar_formato_hd_desde_marcados(
        session=db_session,
        xlsx_intermedio=str(out),
        base_fotos=str(tmp_path),
    )
    assert n == 1  # solo p1 estaba marcado
    assert out.exists()

    import openpyxl
    wb = openpyxl.load_workbook(str(out))
    ws = wb.active
    assert ws.cell(2, 2).value == "A1"
    # Col O = Venta HD MXN (paso 11 redondeado al entero); con fob=10 + pzas=1000
    # el motor produce un valor > 0. Verificamos que sea entero y positivo.
    venta_hd = ws.cell(2, 15).value
    assert isinstance(venta_hd, int) and venta_hd > 0, f"venta_hd inesperado: {venta_hd!r}"
    # Col P = Retail c/IVA MXN, tambien entero > venta_hd
    retail = ws.cell(2, 16).value
    assert isinstance(retail, int) and retail > venta_hd
    # Col Q = Margen Lloyds real (decimal)
    assert isinstance(ws.cell(2, 17).value, float)
    assert len(ws._images) == 1
