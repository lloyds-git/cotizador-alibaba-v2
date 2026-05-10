"""Script one-shot para crear el xlsx de prueba. Correr una vez."""
import openpyxl
from openpyxl.drawing.image import Image
from pathlib import Path

ROOT = Path(__file__).parent
ROOT.mkdir(parents=True, exist_ok=True)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Cotizacion PDF"

headers = ["Foto", "SKU", "Descripcion", "Medidas", "Material", "Peso (kg)",
           "Color", "MOQ", "Packing", "Carton dims", "CBM",
           "Pzas 20ft", "Pzas 40hq", "Lead time", "FOB USD"]
for i, h in enumerate(headers, 1):
    ws.cell(row=1, column=i, value=h)

ws.cell(row=2, column=2, value="TEST-001")
ws.cell(row=2, column=3, value="Producto demo grande")
ws.cell(row=2, column=4, value="L100xW50xH30 mm")
ws.cell(row=2, column=5, value="PP")
ws.cell(row=2, column=6, value="2.5")
ws.cell(row=2, column=8, value="300 pcs")
ws.cell(row=2, column=11, value="0.05")
ws.cell(row=2, column=12, value="500")
ws.cell(row=2, column=15, value=10.5)

ws.cell(row=3, column=2, value="TEST-002")
ws.cell(row=3, column=3, value="Producto demo chico")
ws.cell(row=3, column=15, value=5.0)

# PNG minimo valido 1x1 transparente (67 bytes)
png_bytes = bytes.fromhex(
    "89504e470d0a1a0a0000000d49484452000000010000000108060000001f15c4"
    "890000000a49444154789c6300010000000500010d0a2db40000000049454e44ae426082"
)
foto_path = ROOT / "demo.png"
foto_path.write_bytes(png_bytes)
img = Image(str(foto_path))
img.anchor = "A2"
ws.add_image(img)

wb.save(ROOT / "_intermedio_demo.xlsx")
print(f"Creado: {ROOT / '_intermedio_demo.xlsx'}")
