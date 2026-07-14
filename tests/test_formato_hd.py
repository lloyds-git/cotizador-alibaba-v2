"""Tests del llenado HD con openpyxl (app/formato_hd.py, ruta Linux)."""
import shutil

import openpyxl
import pytest
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

from app import formato_hd


def test_col_letra_a_num():
    assert formato_hd.col_letra_a_num("A") == 1
    assert formato_hd.col_letra_a_num("Z") == 26
    assert formato_hd.col_letra_a_num("AA") == 27


def test_parsear_mapeo():
    m = formato_hd.parsear_mapeo("B=5,C=8,O=11")
    assert m == {2: 5, 3: 8, 15: 11}
    with pytest.raises(ValueError):
        formato_hd.parsear_mapeo("B5")  # falta '='


def _png(path, w=80, h=60, color=(200, 30, 30)):
    PILImage.new("RGB", (w, h), color).save(path)
    return path


def _template_hd(path):
    """Template minimo con el layout HD: labels en A/B, muestra en C."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detalle de productos"
    labels = {
        2: "FOTO", 3: "Vendor Name", 4: "Vendor Number", 5: "SKU", 6: "ITEM TYPE",
        7: "UPC", 8: "Description", 9: "Model", 10: "COUNTRY", 11: "Domestic Cost",
        12: "MOQ Domestic", 13: "DI COST", 14: "MOQ DI", 15: "Pieces per Container",
        16: "Retail", 17: "THD Margin Dom", 18: "THD Margin DI",
    }
    for r, lab in labels.items():
        ws.cell(r, 1, value=lab)
        ws.cell(r, 2, value=lab.upper())
    # Producto de muestra en C (debe quedar sobrescrito)
    ws.cell(8, 3, value="MUESTRA VIEJA")
    ws.cell(17, 3, value="=((C16/1.16)-(C11*19))/(C16/1.16)")
    wb.save(path)
    return path


def _intermedio(path, tmp_path, productos):
    """Intermedio estilo exportar.py: fila 1 headers, datos desde fila 2,
    foto flotante en col A."""
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["Foto", "SKU", "Descripcion", "Medidas", "Material", "Peso", "Color",
               "MOQ", "Packing", "Carton", "CBM", "20ft", "40hq", "Lead", "Venta HD",
               "Retail", "Margen", "Proveedor"]
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, value=h)
    for idx, p in enumerate(productos):
        r = 2 + idx
        ws.cell(r, 2, value=p["sku"])
        ws.cell(r, 3, value=p["desc"])
        ws.cell(r, 8, value=p["moq"])
        ws.cell(r, 13, value=p["pzas40"])
        ws.cell(r, 15, value=p["venta"])
        ws.cell(r, 16, value=p["retail"])
        ws.cell(r, 17, value=p["margen"])
        if p.get("foto"):
            img = XLImage(str(p["foto"]))
            img.anchor = f"A{r}"
            ws.add_image(img)
    wb.save(path)
    return path


def test_llenar_formato_hd_roundtrip(tmp_path):
    template = _template_hd(tmp_path / "template.xlsx")
    foto1 = _png(tmp_path / "f1.png")
    productos = [
        {"sku": "SKU-1", "desc": "Cama para perro", "moq": 200, "pzas40": 2000,
         "venta": 20.44, "retail": 699, "margen": 0.35, "foto": foto1},
        {"sku": "SKU-2", "desc": "Rascador gato", "moq": 100, "pzas40": 1500,
         "venta": 15.0, "retail": 499, "margen": 0.42, "foto": None},
    ]
    origen = _intermedio(tmp_path / "_intermedio_x.xlsx", tmp_path, productos)
    salida = tmp_path / "formato-hd-x.xlsx"

    mapeo = formato_hd.parsear_mapeo("B=5,C=8,H=12,M=15,O=11,P=16,Q=17")
    n = formato_hd.llenar_formato_hd(str(origen), str(template), str(salida), mapeo)
    assert n == 2

    wb = openpyxl.load_workbook(salida)
    ws = wb.active
    # Producto 1 en columna C (3)
    assert ws.cell(5, 3).value == "SKU-1"      # B->5
    assert ws.cell(8, 3).value == "Cama para perro"  # C->8 (sobrescribe MUESTRA VIEJA)
    assert ws.cell(11, 3).value == 20.44       # O->11
    assert ws.cell(16, 3).value == 699         # P->16
    assert ws.cell(17, 3).value == 0.35        # Q->17 (sobrescribe la formula)
    assert ws.cell(9, 3).value == "MPT-0001"   # MPT autoincremental
    assert ws.cell(3, 3).value == "Totikay Pets SA de CV"  # constante
    assert ws.cell(10, 3).value == "China"     # constante
    # Producto 2 en columna D (4)
    assert ws.cell(5, 4).value == "SKU-2"
    assert ws.cell(9, 4).value == "MPT-0002"
    # Formatos numericos aplicados
    assert ws.cell(11, 3).number_format == formato_hd.FMT_MONEDA
    assert ws.cell(17, 3).number_format == formato_hd.FMT_PORCENTAJE
    # Una sola imagen (producto 1 tenia foto; producto 2 no)
    assert len(ws._images) == 1


def test_llenar_sin_productos_lanza(tmp_path):
    template = _template_hd(tmp_path / "t.xlsx")
    origen = _intermedio(tmp_path / "_intermedio_vacio.xlsx", tmp_path, [])
    with pytest.raises(ValueError):
        formato_hd.llenar_formato_hd(
            str(origen), str(template), str(tmp_path / "out.xlsx"),
            formato_hd.parsear_mapeo("B=5,C=8"),
        )


@pytest.mark.skipif(
    formato_hd._soffice_bin() is None, reason="LibreOffice (soffice) no disponible"
)
def test_convertir_a_xlsx(tmp_path):
    # Crea un xlsx, lo convierte a xlsx via LibreOffice (round-trip valido).
    src = tmp_path / "muestra.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "hola"
    wb.save(src)
    out = formato_hd.convertir_a_xlsx(src, tmp_path / "conv")
    assert out.exists() and out.suffix == ".xlsx"
