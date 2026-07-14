from pathlib import Path
import openpyxl
from app.ingest import ingestar_xlsx_intermedio


def _construir_xlsx_minimo(path: Path, filas: list[dict]) -> None:
    """Crea un xlsx intermedio con las 16 columnas del schema actual.

    Cada elemento de `filas` es un dict con las claves (cualquier subset):
    sku, descripcion, fob, cbm, pzas_40hq, pzas_caja.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cotizacion PDF"
    headers = ["Foto", "SKU", "Descripcion", "Medidas", "Material", "Peso (kg)",
               "Color", "MOQ", "Packing", "Carton dims", "CBM",
               "Pzas 20ft", "Pzas 40hq", "Lead time", "FOB USD", "Pzas/caja"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    for r, f in enumerate(filas, start=2):
        ws.cell(row=r, column=2, value=f.get("sku"))
        ws.cell(row=r, column=3, value=f.get("descripcion") or "x")
        ws.cell(row=r, column=11, value=f.get("cbm"))
        ws.cell(row=r, column=13, value=f.get("pzas_40hq"))
        ws.cell(row=r, column=15, value=f.get("fob"))
        ws.cell(row=r, column=16, value=f.get("pzas_caja"))
    wb.save(path)


def test_proveedor_forzado_tiene_precedencia(db_session, tmp_path):
    """proveedor_forzado manda los productos al proveedor elegido (no crea otro)."""
    from app.modelos import Proveedor, Producto
    xlsx = tmp_path / "_intermedio_Goodrich-2.xlsx"
    _construir_xlsx_minimo(xlsx, [{"sku": "56-1", "fob": 2.0}, {"sku": "56-2", "fob": 3.0}])

    ingestar_xlsx_intermedio(
        session=db_session, xlsx_path=str(xlsx),
        nombre_proveedor="Ignorado", fotos_destino=str(tmp_path / "fotos"),
        proveedor_forzado="Tianjin Goodrich",
    )
    db_session.commit()
    provs = [p.nombre for p in db_session.query(Proveedor).all()]
    assert provs == ["Tianjin Goodrich"]
    assert db_session.query(Producto).filter_by(sku="56-1").first().proveedor.nombre == "Tianjin Goodrich"


def test_producto_guarda_archivo_pdf(db_session, tmp_path):
    """Cada producto guarda su PDF de origen (archivo_pdf)."""
    from app.modelos import Producto
    xlsx = tmp_path / "_intermedio_Goodrich-2.xlsx"
    _construir_xlsx_minimo(xlsx, [{"sku": "56-9", "fob": 1.0}])
    ingestar_xlsx_intermedio(
        session=db_session, xlsx_path=str(xlsx),
        nombre_proveedor="V", fotos_destino=str(tmp_path / "fotos"),
    )
    db_session.commit()
    p = db_session.query(Producto).filter_by(sku="56-9").first()
    # Sin meta.json, resolver_archivo_pdf cae al nombre del xlsx intermedio.
    assert p.archivo_pdf == "_intermedio_Goodrich-2.xlsx"


def test_ingest_inserta_proveedor_y_productos(db_session, tmp_path):
    fixture = Path(__file__).parent / "fixtures" / "_intermedio_demo.xlsx"
    fotos_dir = tmp_path / "fotos"

    n = ingestar_xlsx_intermedio(
        session=db_session,
        xlsx_path=str(fixture),
        nombre_proveedor="Demo Vendor",
        fotos_destino=str(fotos_dir),
    )
    db_session.commit()
    assert n == 2

    from app.modelos import Proveedor, Producto
    prov = db_session.query(Proveedor).filter_by(nombre="Demo Vendor").first()
    assert prov is not None
    assert len(prov.productos) == 2

    p1 = db_session.query(Producto).filter_by(sku="TEST-001").first()
    assert p1.fob_usd == 10.5
    assert p1.material == "PP"
    assert p1.cbm == 0.05
    assert p1.pzas_20ft == 500
    assert len(p1.fotos) == 1
    foto_path = fotos_dir / Path(p1.fotos[0].ruta_relativa).name
    assert foto_path.exists()


def test_ingest_idempotente(db_session, tmp_path):
    """Re-ingerir el mismo xlsx no duplica productos."""
    fixture = Path(__file__).parent / "fixtures" / "_intermedio_demo.xlsx"
    fotos_dir = tmp_path / "fotos"

    ingestar_xlsx_intermedio(db_session, str(fixture), "V", str(fotos_dir))
    db_session.commit()
    n2 = ingestar_xlsx_intermedio(db_session, str(fixture), "V", str(fotos_dir))
    db_session.commit()
    assert n2 == 0
    from app.modelos import Producto
    assert db_session.query(Producto).count() == 2


def test_ingest_actualiza_fob_si_cambio(db_session, tmp_path):
    """Si el precio cambio, debe actualizar el registro existente."""
    fixture = Path(__file__).parent / "fixtures" / "_intermedio_demo.xlsx"
    fotos_dir = tmp_path / "fotos"

    ingestar_xlsx_intermedio(db_session, str(fixture), "V", str(fotos_dir))
    db_session.commit()

    from app.modelos import Producto
    p = db_session.query(Producto).filter_by(sku="TEST-001").first()
    p.fob_usd = 99.0
    db_session.commit()

    ingestar_xlsx_intermedio(db_session, str(fixture), "V", str(fotos_dir))
    db_session.commit()

    p2 = db_session.query(Producto).filter_by(sku="TEST-001").first()
    assert p2.fob_usd == 10.5


def test_ingest_persiste_pzas_caja(db_session, tmp_path):
    """pzas_caja en columna 16 se persiste tal cual."""
    xlsx = tmp_path / "t.xlsx"
    _construir_xlsx_minimo(xlsx, [
        {"sku": "PC-001", "fob": 1.0, "pzas_caja": 24, "cbm": 0.05, "pzas_40hq": 100},
    ])
    ingestar_xlsx_intermedio(db_session, str(xlsx), "Vendor", str(tmp_path / "fotos"))
    db_session.commit()

    from app.modelos import Producto
    p = db_session.query(Producto).filter_by(sku="PC-001").first()
    assert p.pzas_caja == 24
    assert p.pzas_40hq == 100  # no se sobrescribe lo leido


def test_ingest_deriva_pzas_40hq_desde_pzas_caja(db_session, tmp_path):
    """Si pzas_40hq viene vacio pero hay pzas_caja+cbm, lo deriva."""
    xlsx = tmp_path / "t.xlsx"
    _construir_xlsx_minimo(xlsx, [
        {"sku": "PC-002", "fob": 1.0, "pzas_caja": 24, "cbm": 0.05},  # sin pzas_40hq
    ])
    ingestar_xlsx_intermedio(db_session, str(xlsx), "Vendor", str(tmp_path / "fotos"))
    db_session.commit()

    from app.modelos import Producto
    p = db_session.query(Producto).filter_by(sku="PC-002").first()
    # floor(67/0.05) * 24 = 1340 * 24 = 32160
    assert p.pzas_40hq == 32160


def test_ingest_no_deriva_sin_pzas_caja(db_session, tmp_path):
    """Sin pzas_caja no se deriva pzas_40hq."""
    xlsx = tmp_path / "t.xlsx"
    _construir_xlsx_minimo(xlsx, [
        {"sku": "PC-003", "fob": 1.0, "cbm": 0.05},
    ])
    ingestar_xlsx_intermedio(db_session, str(xlsx), "Vendor", str(tmp_path / "fotos"))
    db_session.commit()

    from app.modelos import Producto
    p = db_session.query(Producto).filter_by(sku="PC-003").first()
    assert p.pzas_caja is None
    assert p.pzas_40hq is None
